#!/usr/bin/env python3
"""
cards.py — 詞彙卡片語料前置生成工具
輸入主題 → OpenAI 生成 → 輸出 cards/{topic}.xlsx
已做過的主題自動跳過，跨主題詞彙自動去重。
"""

import glob
import json
import math
import os
import random
import re
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openai import OpenAI
from dotenv import load_dotenv

load_dotenv()

BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
CARDS_DIR       = os.path.join(BASE_DIR, "cards")
OUTPUT_DIR      = os.path.join(BASE_DIR, "output")
USED_WORDS_FILE = os.path.join(BASE_DIR, "used_words.json")
GENERATE_CHUNK  = 20

os.makedirs(CARDS_DIR,  exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

OPENAI_KEYS = [k for k in [os.getenv("OPENAI_API_KEY"), os.getenv("OPENAI_API_KEY_2")] if k]

HEADERS = ["id", "word_en", "word_ipa", "word_cn", "tips",
           "sentence_en", "sentence_ipa", "sentence_cn"]

FIELD_SPEC = """Return a JSON object with a single key "items" whose value is an array of objects.
Each object MUST have exactly these keys:
- "word_en"     : the English word or common phrase
- "word_ipa"    : IPA pronunciation of the word/phrase, enclosed in forward slashes (e.g., "/tʃɑp ˈvɛdʒtəblz/")
- "word_cn"     : Traditional Chinese translation (繁體中文)
- "tips"        : 極短一句話，不要 emoji。提供記憶法、字根拆解、常見搭配詞或使用情境提示。禁止重複 word_cn 的中文解釋。
- "sentence_en" : a natural, conversational English example sentence (not textbook style)
- "sentence_ipa": full IPA pronunciation of the example sentence, enclosed in forward slashes (e.g., "/kæn juː hɛlp miː.../")
- "sentence_cn" : 台灣繁體中文口語意譯（非逐字翻譯）"""


def _call_openai(messages: list, **kwargs):
    from openai import RateLimitError, AuthenticationError, APIError
    if not OPENAI_KEYS:
        raise RuntimeError("未設定任何 OPENAI_API_KEY，請在 .env 補上金鑰")
    last_err = None
    for idx, key in enumerate(OPENAI_KEYS):
        try:
            client = OpenAI(api_key=key)
            return client.chat.completions.create(messages=messages, **kwargs)
        except (RateLimitError, AuthenticationError) as e:
            print(f"   ⚠️  OpenAI 金鑰 #{idx + 1} 無法使用（{type(e).__name__}），切換備用金鑰...")
            last_err = e
        except APIError as e:
            last_err = e
    raise last_err


def _normalize_key(word: str) -> str:
    if not word:
        return ""
    # Lowercase, strip punctuation, remove articles (a, an, the) to prevent duplicates like "Season meat" vs "Season the meat"
    w = word.lower().strip()
    w = re.sub(r'[.,!?\-_\'"]', '', w)
    w = re.sub(r'\b(the|a|an)\b', '', w)
    return ' '.join(w.split())


def _is_valid_item(item: dict) -> bool:
    required_keys = ["word_en", "word_ipa", "word_cn", "tips", "sentence_en", "sentence_ipa", "sentence_cn"]
    for key in required_keys:
        val = item.get(key)
        if not val or not isinstance(val, str) or not val.strip():
            return False
    
    # Check if word_ipa or sentence_ipa is just the English text (failed to generate IPA)
    word_en_clean = _normalize_key(item["word_en"])
    word_ipa_clean = _normalize_key(item["word_ipa"])
    if word_en_clean == word_ipa_clean:
        return False
        
    sentence_en_clean = _normalize_key(item["sentence_en"])
    sentence_ipa_clean = _normalize_key(item["sentence_ipa"])
    if sentence_en_clean == sentence_ipa_clean:
        return False
        
    if len(item["word_ipa"].strip()) < 2 or len(item["sentence_ipa"].strip()) < 2:
        return False
        
    return True


def _build_prompt(topic: str, count: int) -> str:
    return f"""以「{topic}」實用的對話為主題，盡量涵蓋各種「{topic}」的狀況，給我 {count} 句。

規則：
1. 以實用對話情境為核心，每個詞彙都要是在「{topic}」場景中真正會用到的。
2. tips 極度短，一句話就好，不要 emoji，不要重複中文解釋。用記憶法、字根拆解或常見搭配詞皆可。
3. sentence_en 必須是實際對話中會說的自然句子，不要教科書式的生硬句子。
4. 所有中文（word_cn, tips, sentence_cn）使用台灣繁體中文日常口語，不要逐字翻譯。
5. 涵蓋「{topic}」的各種不同子情境，不要集中在同一個場景。
6. 優先收錄片語、搭配詞、慣用語，而非單一簡單單字。
7. 必須為每個詞彙和句子生成精確且完整的 IPA 國際音標。國際音標必須使用斜線「/」包裹（例如：/ˈæp.əl/）。絕對不能遺漏音標欄位，也絕對不能直接填入英文拼寫。
"""


def _load_used_words() -> set[str]:
    if os.path.exists(USED_WORDS_FILE):
        try:
            with open(USED_WORDS_FILE, "r", encoding="utf-8") as f:
                return set(json.load(f))
        except Exception:
            pass
    return set()


def _save_used_words(words: set[str]):
    tmp = USED_WORDS_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(sorted(words), f, ensure_ascii=False, indent=2)
    os.replace(tmp, USED_WORDS_FILE)


def _existing_topics() -> list[str]:
    files = glob.glob(os.path.join(CARDS_DIR, "*.xlsx"))
    return [os.path.splitext(os.path.basename(f))[0] for f in sorted(files)]


def _topic_to_slug(topic: str) -> str:
    return topic.strip().replace(" ", "_").replace("-", "_")


def generate(topic: str, count: int = 50) -> list[dict]:
    prompt = _build_prompt(topic, count)
    used_words = _load_used_words()
    seen_normalized = { _normalize_key(w) for w in used_words if w }
    all_items: list = []
    # Increase max_rounds in case some generated items are filtered out/invalid
    max_rounds = math.ceil(count / GENERATE_CHUNK) * 2 + 5
    rounds = 0
    consecutive_fails = 0

    while len(all_items) < count and rounds < max_rounds:
        need = count - len(all_items)
        chunk_size = min(GENERATE_CHUNK, need + 3)
        print(f"   🔄 OpenAI #{rounds + 1}（目標 {chunk_size} 個，已有 {len(all_items)} 個）...")

        # Combine current session words and a sample of historical used words to prevent duplicates
        current_words = [item["word_en"] for item in all_items if item.get("word_en")]
        exclude_set = set(current_words)
        if used_words:
            sample_size = min(400, len(used_words))
            exclude_set.update(random.sample(list(used_words), sample_size))

        exclusion_note = ""
        if exclude_set:
            exclusion_note = f"\nNEVER generate any of these already-used words/phrases: {', '.join(sorted(exclude_set))}\n"

        full_prompt = prompt + exclusion_note + f"\nGenerate exactly {chunk_size} items.\n" + FIELD_SPEC
        try:
            resp = _call_openai(
                messages=[{"role": "user", "content": full_prompt}],
                model="gpt-4o-mini",
                response_format={"type": "json_object"},
                # Lower temperature to 0.4 for higher quality, structure compliance, and less random hallucinations
                temperature=0.4,
            )
            raw = json.loads(resp.choices[0].message.content).get("items", [])
        except Exception as e:
            print(f"      ⚠️ API 呼召失敗，等待 2 秒後重試: {e}")
            time.sleep(2)
            consecutive_fails += 1
            if consecutive_fails >= 3:
                print("   ⚠️  連續失敗 3 次，中斷生成。")
                break
            continue

        consecutive_fails = 0
        for item in raw:
            if not _is_valid_item(item):
                print(f"      ⚠️ 跳過格式或音標無效的項目: {item.get('word_en', 'Unknown')}")
                continue
            
            raw_word = item.get("word_en", "")
            key = _normalize_key(raw_word)
            if key and key not in seen_normalized:
                w = raw_word.strip()
                item["word_en"] = w[0].upper() + w[1:] if w else w
                
                # Format IPA columns to always have forward slashes
                for ipa_key in ["word_ipa", "sentence_ipa"]:
                    ipa_val = item[ipa_key].strip()
                    if not ipa_val.startswith("/"):
                        ipa_val = "/" + ipa_val
                    if not ipa_val.endswith("/"):
                        ipa_val = ipa_val + "/"
                    item[ipa_key] = ipa_val
                
                all_items.append(item)
                seen_normalized.add(key)
        rounds += 1

    all_items = all_items[:count]
    for i, item in enumerate(all_items):
        item["id"] = f"{i + 1:02d}"

    new_keys = {item["word_en"].lower().strip().strip(".,!?") for item in all_items if item.get("word_en")}
    _save_used_words(used_words | new_keys)
    return all_items


def _chapter_time(seconds: float) -> str:
    m = int(seconds) // 60
    s = int(seconds) % 60
    return f"{m:02d}:{s:02d}"


def _parse_srt_starts(srt_path: str) -> list[float]:
    """讀取 SRT，回傳每條字幕的起始秒數。"""
    if not os.path.exists(srt_path):
        return []
    starts: list[float] = []
    with open(srt_path, "r", encoding="utf-8") as f:
        for line in f:
            m = re.match(r"(\d{2}):(\d{2}):(\d{2})[,.](\d{3})\s*-->", line)
            if m:
                h, mm, ss, ms = (int(x) for x in m.groups())
                starts.append(h * 3600 + mm * 60 + ss + ms / 1000.0)
    return starts


def _generate_yt_title(topic: str) -> str:
    prompt = (
        f"你是台灣 YouTube 英語教學頻道的標題撰稿人。"
        f"請為主題「{topic}」寫一句 YouTube 影片標題（繁體中文，25-40 字）。"
        f"風格：吸睛、有痛點、含具體場景與 Rayo 智慧閃卡關鍵字。"
        f"只輸出標題本身，不要引號、不要 hashtag、不要多行。"
    )
    try:
        resp = _call_openai(
            messages=[{"role": "user", "content": prompt}],
            model="gpt-4o-mini",
            temperature=0.9,
            max_tokens=80,
        )
        title = resp.choices[0].message.content.strip().strip('「」""\'\'')
        return title.splitlines()[0] if title else ""
    except Exception as e:
        print(f"⚠️  OpenAI 生成 YouTube 標題失敗 ({e})，使用預設模板")
        return f"【日常英文】{topic} 英文懶人包｜Rayo 智慧閃卡陪你 14 天上手"


def _generate_yt_topic_paragraph(topic: str) -> str:
    prompt = (
        f"你是一位台灣 YouTube 英語教學頻道的文案寫手。"
        f"請為主題「{topic}」寫一段 YouTube 影片描述（約 100-150 字繁體中文）。"
        f"格式要求：**第一句必須是一個以問號結尾的痛點/情境 hook**，"
        f"例如「暑假馬上就要飛了，卻發現英文還沒準備好？」的風格，"
        f"貼合「{topic}」情境。"
        f"接著的 2-3 句延續 hook，介紹本集內容（濃縮的實用情境／金句），"
        f"並明確提到搭配 Rayo 智慧閃卡與影子跟讀（Shadowing）練習。"
        f"風格活潑、有吸引力、貼近台灣觀眾口吻。"
        f"不要加標題、不要加 hashtag、不要加連結、不要換行，全部合成一段。"
    )
    try:
        resp = _call_openai(
            messages=[{"role": "user", "content": prompt}],
            model="gpt-4o-mini",
            temperature=0.9,
            max_tokens=400,
        )
        return resp.choices[0].message.content.strip()
    except Exception as e:
        print(f"⚠️  OpenAI 生成 YouTube 描述失敗 ({e})，使用預設模板")
        return (
            f"想學「{topic}」實用英文卻不知從何開始？"
            f"這集為你準備了最實用的「{topic}」英文懶人包，"
            f"專為沒時間準備的零基礎新手設計。"
            f"搭配 Rayo 智慧閃卡與影子跟讀（Shadowing），"
            f"每天只需幾分鐘，把最實用的金句印在腦海裡！"
        )


def _generate_yt_hashtags(topic: str) -> list[str]:
    prompt = (
        f"為 YouTube 英語教學影片主題「{topic}」生成 8-12 個 SEO 標籤，"
        f"涵蓋：繁體中文（如「{topic}英文、{topic}單字」等 2-4 個同義詞）、"
        f"簡體中文（2-3 個）、英文小寫（3-5 個，如「kitchen english、cooking vocabulary」風格）。"
        f'只輸出 JSON：{{"tags": ["tag1", "tag2", ...]}}，不要多餘文字。'
    )
    try:
        resp = _call_openai(
            messages=[{"role": "user", "content": prompt}],
            model="gpt-4o-mini",
            temperature=0.7,
            max_tokens=300,
            response_format={"type": "json_object"},
        )
        data = json.loads(resp.choices[0].message.content)
        tags = data.get("tags", [])
        cleaned = [t.strip() for t in tags if isinstance(t, str) and t.strip()]
        if cleaned:
            return cleaned
    except Exception as e:
        print(f"⚠️  OpenAI 生成主題 hashtags 失敗 ({e})，使用預設 fallback")
    return [f"{topic}英文", f"{topic} english"]


def _count_xlsx_rows(xlsx_path: str) -> int:
    """讀取現有 xlsx 的資料筆數（扣除 header 列）。"""
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        return max(0, ws.max_row - 1)
    except Exception:
        return 0


def write_youtube_description(topic: str, card_count: int, output_path: str):
    """產出 youtube_{topic}.txt。若對應的 SRT 已存在，四個進度時間戳從中讀取；否則用 00:00。"""
    slug = _topic_to_slug(topic)
    srt_path = os.path.join(OUTPUT_DIR, f"final_{slug.lower()}.srt")
    srt_starts = _parse_srt_starts(srt_path)

    def _srt_time(idx: int) -> str:
        if 0 <= idx < len(srt_starts):
            return _chapter_time(srt_starts[idx])
        return "00:00"

    ts_start = "00:00"
    if srt_starts:
        ts_25 = _srt_time(25)
        ts_50 = _srt_time(card_count)
        ts_75 = _srt_time(card_count + 25)
        print(f"📼 已找到 SRT，四個進度時間戳從字幕讀取")
    else:
        ts_25 = ts_50 = ts_75 = "00:00"
        print(f"ℹ️  未找到 {srt_path}，時間戳先用 00:00 佔位（跑完影片後可重新產生）")

    title = _generate_yt_title(topic)
    paragraph = _generate_yt_topic_paragraph(topic)
    topic_tags = _generate_yt_hashtags(topic)

    fixed_hashtags = ["英文學習", "日常對話", "14天挑戰", "英文口說", "影子跟讀", "英語教學"]
    hashtag_line = " ".join(f"#{t}" for t in fixed_hashtags)

    extended_tags = [
        "Rayo智慧閃卡", "間隔重複", "Shadowing",
        "英文學習", "英语学习", "learn english",
        "零基礎英文", "生活英文", "daily english", "english speaking practice",
    ]
    seen: set[str] = set()
    combined_tags: list[str] = []
    for t in extended_tags + list(topic_tags):
        key = t.lower()
        if key in seen:
            continue
        seen.add(key)
        combined_tags.append(t)
    comma_line = ", ".join(combined_tags)

    lines = [
        title,
        "",
        paragraph,
        "",
        "👇 搭配 Rayo 智慧閃卡，學習效率翻倍 👇",
        "官網：https://rayo-ai.com/",
        "iOS App：https://rayo.pse.is/8ugjnq",
        "Chrome 插件：https://rayo.pse.is/8ughfh",
        "",
        f"{ts_start} 開始學習！",
        f"{ts_25} 25%繼續加油！",
        f"{ts_50} 50% 再複習一次  GO! GO!",
        f"{ts_75} 75% 最後衝刺！",
        "",
        "✅ 訂閱頻道並開啟小鈴鐺",
        "💬 在下方留言告訴我：你覺得最難開口的一句英文是什麼？",
        "",
        hashtag_line,
        "",
        comma_line,
    ]
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"✅ YouTube 描述: {output_path}")


def write_xlsx(items: list[dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cards"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        ws.append([item.get(h, "") for h in HEADERS])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(path)


def main():
    existing = _existing_topics()
    if existing:
        print(f"✅ 已有主題: {', '.join(existing)}")

    topic = input("\n📌 請輸入主題名稱: ").strip()
    if not topic:
        print("⛔ 主題不能為空")
        return

    slug = _topic_to_slug(topic)
    xlsx_path    = os.path.join(CARDS_DIR, f"{slug}.xlsx")
    yt_desc_path = os.path.join(CARDS_DIR, f"youtube_{slug}.txt")

    xlsx_exists    = os.path.exists(xlsx_path)
    yt_desc_exists = os.path.exists(yt_desc_path)

    if xlsx_exists and yt_desc_exists:
        print(f"⚠️  「{topic}」的詞卡與 YouTube 描述都已存在，全部跳過")
        return

    if xlsx_exists:
        card_count = _count_xlsx_rows(xlsx_path) or 50
        print(f"📄 「{topic}」詞卡已存在（{card_count} 個詞彙），只補產 YouTube 描述")
        write_youtube_description(topic, card_count, yt_desc_path)
        return

    raw_count = input("🔢 卡片數量（留空=50）: ").strip()
    count = int(raw_count) if raw_count.isdigit() and int(raw_count) > 0 else 50

    used_before = len(_load_used_words())
    print(f"\n🆕 開始生成「{topic}」({count} 個詞彙)...")
    items = generate(topic, count)

    if not items:
        print("❌ 未生成任何詞彙")
        return

    write_xlsx(items, xlsx_path)
    used_after = len(_load_used_words())
    print(f"\n✅ 已生成 {len(items)} 個詞彙 → {xlsx_path}")
    print(f"📝 used_words.json 已更新（{used_before} → {used_after}）")

    write_youtube_description(topic, len(items), yt_desc_path)


if __name__ == "__main__":
    main()
