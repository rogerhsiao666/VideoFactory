#!/usr/bin/env python3
import sys
print("=== 環境診斷 ===")
print(f"Python 路徑: {sys.executable}")
print(f"Python 版本: {sys.version}")
try:
    import youtube_transcript_api as _yta
    print(f"套件路徑: {_yta.__file__}")
    from importlib.metadata import version as _pkg_version
    try:
        print(f"套件版本: {_pkg_version('youtube-transcript-api')}")
    except Exception:
        print(f"套件版本: {getattr(_yta, '__version__', '版本未知')}")
    from youtube_transcript_api import YouTubeTranscriptApi
    print(f"YouTubeTranscriptApi 方法: {[m for m in dir(YouTubeTranscriptApi) if not m.startswith('_')]}")
except ImportError as e:
    print(f"套件未安裝: {e}")
print("=== 診斷結束 ===")

"""
VideoFactory Enterprise Edition
Automated YouTube Content Creation Pipeline
Topic → OpenAI → Pexels → TTS → Pillow → MoviePy → Final MP4 + SRT + Description
"""

import glob
import json
import math
import os
import re
import termios
import textwrap
import asyncio
import random
import time
import gc
import shutil
import subprocess
import requests
import edge_tts
from openai import OpenAI
from PIL import Image, ImageDraw, ImageFont, ImageFilter
from moviepy import ImageClip, AudioFileClip, VideoFileClip, concatenate_videoclips, CompositeAudioClip
from moviepy.audio.AudioClip import AudioClip, concatenate_audioclips
from dotenv import load_dotenv

load_dotenv()

# ── Whisper 模型（karaoke 字幕用）────────────────────
try:
    from faster_whisper import WhisperModel as _WhisperModel
    whisper_model = _WhisperModel("base", device="cpu", compute_type="int8")
except Exception as _e:
    print(f"⚠️ Whisper 模型載入失敗，karaoke 字幕將停用: {_e}")
    whisper_model = None

# ================= 配置區 =================
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(BASE_DIR, "assets")
TEMP_DIR   = os.path.join(BASE_DIR, "temp")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
IMAGES_DIR = os.path.join(TEMP_DIR, "images")
DATA_FILE  = os.path.join(BASE_DIR, "data.json")

os.makedirs(TEMP_DIR,   exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(IMAGES_DIR, exist_ok=True)

FONT_EN     = os.path.join(ASSETS_DIR, "font_en.ttf")
FONT_CN     = os.path.join(ASSETS_DIR, "font_cn.otf")
BGM_DIR     = os.path.join(ASSETS_DIR, "bgm")        # 資料夾：多首 BGM 隨機挑選
BGM_SINGLE  = os.path.join(ASSETS_DIR, "bgm.mp3")    # 備用：單一檔案
MOCKUP_DIR   = os.path.join(ASSETS_DIR, "mockup")
MOTION_DIR   = os.path.join(ASSETS_DIR, "motion")
AUDIO_DIR    = os.path.join(ASSETS_DIR, "Audio")
PREBUILT_DIR = os.path.join(ASSETS_DIR, "prebuilt")


def _pick_prebuilt(segment: str) -> str:
    import glob as _glob
    candidates = sorted(_glob.glob(os.path.join(PREBUILT_DIR, f"{segment}_*.mp4")))
    if candidates:
        return random.choice(candidates)
    # fallback: 舊靜態檔案
    return os.path.join(ASSETS_DIR, f"{segment}.mp4")


INTRO_VIDEO = _pick_prebuilt("intro")
BREAK_VIDEO = _pick_prebuilt("break")
OUTRO_VIDEO = _pick_prebuilt("outro")


async def generate_custom_intro(text: str, voice: str = "zh-TW-HsiaoChenNeural") -> str | None:
    """
    將使用者輸入的文字轉成 TTS mp3，讓使用者聆聽確認後，
    才將音訊疊加到 prebuilt intro 影片上，回傳最終 mp4 路徑；
    取消則回傳 None。支援重試：每次重試獨立 mp3 檔名。
    """
    base_intro = _pick_prebuilt("intro")
    if not os.path.exists(base_intro):
        print("⚠️  找不到 prebuilt intro 影片，無法生成客製化片頭")
        return None

    attempt = 0
    while True:
        attempt += 1
        tts_path = os.path.join(TEMP_DIR, f"custom_intro_tts_{attempt}.mp3")

        print("🎙️  正在生成片頭語音...")
        await generate_audio(text, voice, tts_path)

        # 直接播放 mp3 預覽（不合成影片）
        print(f"\n▶  預覽片頭音訊：{tts_path}")
        try:
            if sys.platform == "win32":
                os.startfile(tts_path)
            elif sys.platform == "darwin":
                subprocess.run(["open", tts_path], check=True)
            else:
                subprocess.run(["xdg-open", tts_path], check=True)
        except Exception:
            print("   (無法自動開啟播放器，請手動點擊上方路徑試聽)")

        confirm = input("✅  確認使用這個片頭音訊？(y=確認使用 / r=重新嘗試 / n=使用預設，預設 y): ").strip().lower()
        if confirm in ("", "y"):
            # 確認後才合成影片
            final_path = os.path.join(TEMP_DIR, f"custom_intro_final_{attempt}.mp4")
            print("🎬  正在合成片頭影片（僅在確認後執行一次）...")
            tts_clip = _load_audio(tts_path)
            tts_dur  = tts_clip.duration + 0.5
            tts_clip.close()
            cmd = [
                "ffmpeg", "-y", "-stream_loop", "-1", "-i", base_intro, "-i", tts_path,
                "-c:v", "libx264", "-preset", "ultrafast", "-c:a", "aac",
                "-map", "0:v:0", "-map", "1:a:0", "-t", str(round(tts_dur, 3)),
                final_path, "-loglevel", "error"
            ]
            ret = subprocess.run(cmd)
            if ret.returncode != 0 or not os.path.exists(final_path):
                print("❌  片頭合成失敗")
                return None
            return final_path
        elif confirm == "r":
            while True:
                new_text = input("   重新輸入片頭文字：").strip()
                if new_text:
                    text = new_text
                    break
                print("   ⚠️ 文字不能為空，請重新輸入。")
        else:
            return None

def create_rounded_highlight_video(
    image_path: str,
    audio_path: str,
    output_mp4: str,
    font_path: str,
    target_word: str = "",
    sentence_en: str = "",
) -> str | None:
    """
    在靜態卡片圖上疊動態圓角底框，生成 karaoke 影片。
    - 位置：從 sentence_en 原句計算，與靜態卡片像素對齊（textbbox 同 draw_text_with_highlight）
    - 視覺：先畫暗色圓角底框，再把當前詞重繪在框上（框在文字下方）
    - 時序：加 pre-frame 處理 leading silence，最後一幀用實際音訊時長
    """
    if not whisper_model:
        print("   ❌ 缺少 Whisper 模型，無法生成動態字幕")
        return None

    print(f"   🎙️ Whisper 分析音軌: {os.path.basename(audio_path)}")
    segments, _ = whisper_model.transcribe(audio_path, word_timestamps=True)
    words_data = []
    for segment in segments:
        for word in segment.words:
            words_data.append({"word": word.word.strip(), "start": word.start, "end": word.end})

    if not words_data:
        print("   ⚠️ Whisper 找不到任何單字，跳過 karaoke")
        return None

    # 排版參數（對齊靜態卡片 create_sentence_card_image）
    font_size = 80
    start_x   = 150
    start_y   = 300
    max_width = 1600
    padding   = 15
    try:
        font = ImageFont.truetype(font_path, font_size)
    except Exception:
        font = ImageFont.load_default()

    # 目標詞集合（用於判斷橘黃色）
    clean_target  = re.sub(r'[^\w\s]', '', target_word).lower()
    target_tokens = {_normalize_token(t) for t in clean_target.split() if t}

    # ── 計算單字座標（用 textbbox 對齊靜態卡片）──────────────────────
    measure_img  = Image.new("RGBA", (1, 1))
    measure_draw = ImageDraw.Draw(measure_img)
    sentence_words = sentence_en.split() if sentence_en else [w["word"] for w in words_data]

    word_layout = []
    cur_x, cur_y = start_x, start_y
    for word in sentence_words:
        bbox_sp = measure_draw.textbbox((0, 0), word + " ", font=font)
        bbox_w  = measure_draw.textbbox((0, 0), word,       font=font)
        adv_w   = bbox_sp[2] - bbox_sp[0]
        word_h  = bbox_sp[3] - bbox_sp[1]
        if cur_x + adv_w > start_x + max_width:
            cur_x  = start_x
            cur_y += word_h + 20   # line_spacing=20，同靜態卡
        word_layout.append({
            "text":   word,
            "x":      cur_x,   "y":      cur_y,
            "box_x1": cur_x + bbox_w[0], "box_y1": cur_y + bbox_w[1],
            "box_x2": cur_x + bbox_w[2], "box_y2": cur_y + bbox_w[3],
        })
        cur_x += adv_w

    # ── Whisper 時間戳對應原句詞（貪婪字串匹配，解決縮寫斷詞問題）──────
    w_idx = 0
    for layout_item in word_layout:
        target_clean = re.sub(r'[^a-z0-9]', '', layout_item["text"].lower())
        start_time, end_time = None, None
        accumulated_str = ""
        while w_idx < len(words_data):
            w_item = words_data[w_idx]
            if start_time is None:
                start_time = w_item["start"]
            end_time = w_item["end"]
            w_clean = re.sub(r'[^a-z0-9]', '', w_item["word"].lower())
            accumulated_str += w_clean
            w_idx += 1
            if target_clean and (target_clean in accumulated_str or accumulated_str in target_clean):
                break
        layout_item["start"] = start_time if start_time is not None else 0.0
        layout_item["end"]   = end_time   if end_time   is not None else 0.0

    if not word_layout:
        print("   ⚠️ 無法對齊詞彙，跳過 karaoke")
        return None

    # ── 取得音訊總時長（ffprobe）────────────────────────────────────
    probe = subprocess.run(
        ["ffprobe", "-v", "error", "-show_entries", "format=duration",
         "-of", "default=noprint_wrappers=1:nokey=1", audio_path],
        capture_output=True, text=True,
    )
    audio_dur = float(probe.stdout.strip()) if probe.returncode == 0 and probe.stdout.strip() else 10.0

    # ── 逐幀產生 PNG ───────────────────────────────────────────────
    stem       = os.path.splitext(os.path.basename(output_mp4))[0]
    frames_dir = os.path.join(TEMP_DIR, f"kframes_{stem}")
    os.makedirs(frames_dir, exist_ok=True)
    concat_lines: list[str] = []
    base_img = Image.open(image_path).convert("RGBA")

    # pre-frame：leading silence 期間顯示靜態卡片（無框）
    lead_silence = word_layout[0]["start"]
    if lead_silence > 0.05:
        pre_path = os.path.join(frames_dir, "frame_pre.png")
        base_img.convert("RGB").save(pre_path)
        concat_lines.append(f"file '{pre_path.replace(chr(92), '/')}'")
        concat_lines.append(f"duration {lead_silence:.3f}")

    for i, focus in enumerate(word_layout):
        img  = base_img.copy()
        draw = ImageDraw.Draw(img)

        # 1. 先畫暗色圓角底框（框在文字下方）
        draw.rounded_rectangle(
            [focus["box_x1"] - padding, focus["box_y1"] - padding,
             focus["box_x2"] + padding, focus["box_y2"] + padding],
            radius=16, fill=(0, 0, 0, 200),
        )
        # 2. 把當前詞重繪在框上（確保文字在框上方可見）
        is_target  = bool(target_tokens and _normalize_token(focus["text"]) in target_tokens)
        word_color = (255, 210, 0, 255) if is_target else (255, 255, 255, 255)
        draw.text((focus["x"], focus["y"]), focus["text"], font=font, fill=word_color)

        frame_path = os.path.join(frames_dir, f"frame_{i:04d}.png")
        img.convert("RGB").save(frame_path)

        if i + 1 < len(word_layout):
            duration = word_layout[i + 1]["start"] - focus["start"]
        else:
            duration = max(audio_dur - focus["start"] + 0.8, 0.5)
        safe_path = frame_path.replace("\\", "/")
        concat_lines.append(f"file '{safe_path}'")
        concat_lines.append(f"duration {max(duration, 0.05):.3f}")

    concat_file = os.path.join(frames_dir, "concat.txt")
    with open(concat_file, "w", encoding="utf-8") as _cf:
        _cf.write("\n".join(concat_lines))

    print("   🎬 FFmpeg 合成 karaoke 字幕...")
    cmd = [
        "ffmpeg", "-y",
        "-i", audio_path,
        "-f", "concat", "-safe", "0", "-i", concat_file.replace("\\", "/"),
        "-map", "1:v", "-map", "0:a",
        "-c:v", "libx264", "-pix_fmt", "yuv420p",
        "-c:a", "aac", "-b:a", "192k",
        "-shortest",
        output_mp4,
        "-loglevel", "error",
    ]
    ret = subprocess.run(cmd)
    if ret.returncode != 0 or not os.path.exists(output_mp4):
        print("   ❌ karaoke 合成失敗")
        return None
    return output_mp4


BRAND_NAME       = "Rayo: AI Flashcards"
BATCH_SIZE       = 5
GENERATE_CHUNK   = 20          # 每次呼叫 OpenAI 最多要求幾個詞彙
USED_WORDS_FILE  = os.path.join(BASE_DIR, "used_words.json")
BUFFER_TIME      = 0.5   # seconds of silence padding after each audio clip
RECALL_THINK_TIME = 2.0  # 遮中文卡念完後的靜音思考秒數（Phase 2）
FPS = 24


def _load_audio(path: str) -> AudioFileClip:
    """載入音訊。MP3 先轉 WAV 以確保 duration 精確（WAV header 含精確 sample count）。"""
    if path.endswith('.mp3'):
        wav_path = path[:-4] + '.wav'
        subprocess.run(
            ['ffmpeg', '-y', '-i', path, wav_path],
            check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        path = wav_path
    return AudioFileClip(path)


def _image_clip(img_path: str, audio: AudioFileClip, extra_dur: float = 0.0) -> ImageClip:
    """畫面時長 = 音訊時長，聲音是唯一主軸。"""
    if extra_dur > 0.001:
        nch = getattr(audio, 'nchannels', 2)
        fps = getattr(audio, 'fps', 44100)
        make_frame = (lambda t: [0, 0]) if nch == 2 else (lambda t: 0)
        silence = AudioClip(make_frame, duration=extra_dur, fps=fps)
        audio = concatenate_audioclips([audio, silence])
    return ImageClip(img_path).with_duration(audio.duration).with_audio(audio)

PEXELS_KEY   = os.getenv("PEXELS_API_KEY")

# 支援雙金鑰自動輪換：在 .env 設定 OPENAI_API_KEY 與 OPENAI_API_KEY_2
OPENAI_KEYS  = [k for k in [os.getenv("OPENAI_API_KEY"), os.getenv("OPENAI_API_KEY_2")] if k]

# 支援雙金鑰自動輪換：在 .env 設定 TAVILY_API_KEY 與 TAVILY_API_KEY_2
TAVILY_KEYS  = [k for k in [os.getenv("TAVILY_API_KEY"), os.getenv("TAVILY_API_KEY_2")] if k]


# ================= API 功能 =================

def _call_openai(messages: list, **kwargs):
    """
    依序嘗試所有 OPENAI_KEYS，遇到配額/認證錯誤自動切換到下一組金鑰。
    所有金鑰均失敗時拋出最後一個例外。
    金鑰值不會出現在任何輸出或日誌中。
    """
    from openai import RateLimitError, AuthenticationError, APIError
    if not OPENAI_KEYS:
        raise RuntimeError("未設定任何 OPENAI_API_KEY，請在 .env 補上金鑰")

    last_err = None
    for idx, key in enumerate(OPENAI_KEYS):
        try:
            client = OpenAI(api_key=key)
            return client.chat.completions.create(messages=messages, **kwargs)
        except (RateLimitError, AuthenticationError) as e:
            print(f"   ⚠️  OpenAI 金鑰 #{idx + 1} 無法使用（{type(e).__name__}），切換備用金鑰...")
            last_err = e
        except APIError as e:
            # 非配額問題（如網路錯誤）直接拋出，不輪換
            raise
    raise last_err


def _call_tavily(query: str, **kwargs):
    """
    依序嘗試所有 TAVILY_KEYS，遇到配額/認證錯誤自動切換到下一組金鑰。
    回傳 (context_str, key_index_used)。金鑰值不出現在任何輸出中。
    """
    from tavily import TavilyClient
    if not TAVILY_KEYS:
        raise RuntimeError("未設定任何 TAVILY_API_KEY，請在 .env 補上金鑰")

    max_tokens = kwargs.pop("max_tokens", None)

    last_err = None
    for idx, key in enumerate(TAVILY_KEYS):
        try:
            client   = TavilyClient(api_key=key)
            response = client.search(query=query, **kwargs)
            result   = "\n\n".join(r.get("content", "") for r in response.get("results", []))
            if max_tokens:
                result = result[:max_tokens]
            return result, idx + 1
        except Exception as e:
            err_lower = str(e).lower()
            if any(kw in err_lower for kw in ("quota", "limit", "unauthorized", "invalid", "forbidden")):
                print(f"   ⚠️  Tavily 金鑰 #{idx + 1} 無法使用，切換備用金鑰...")
                last_err = e
            else:
                raise
    raise last_err


def export_review_excel(data_list: list, topic: str) -> str:
    """將 data_list 匯出為審核用 Excel，欄位對應卡片所有欄位。回傳檔案路徑。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review"

    headers = ["id", "word_en", "word_ipa", "word_cn", "tips",
               "sentence_en", "sentence_ipa", "sentence_cn"]

    # 標題列格式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 資料列
    for item in data_list:
        ws.append([item.get(h, "") for h in headers])

    # 自動欄寬（最大 60）
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    safe_topic = re.sub(r'[^\w\u4e00-\u9fff-]', '_', topic)
    path = os.path.join(OUTPUT_DIR, f"review_{safe_topic}.xlsx")
    wb.save(path)
    return path


def import_review_excel(path: str) -> list:
    """從審核 Excel 讀回 data_list，保留使用者在 Excel 中做的任何修改。"""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
        if item.get("word_en"):   # 空列跳過
            result.append(item)
    return result


def generate_content(topic: str, count: int, context: str = "") -> list:
    """透過 OpenAI GPT-4o-mini 生成英語詞彙卡片。
    雙引擎邏輯：
      有 context → 引擎 A：Master Prompt，進階知識/新聞單字
      無 context → 引擎 B：Everyday Prompt，生活實用對話單字
    """
    # 欄位規格：下游卡片繪製與 Firestore 匯出均依賴此命名，請勿更動
    field_spec = """Return a JSON object with a single key "items" whose value is an array of objects.
Each object MUST have exactly these keys:
- "word_en"     : the English word or common phrase
- "word_ipa"    : IPA pronunciation of the word/phrase
- "word_cn"     : Traditional Chinese translation
- "tips"        : 提供記憶法、字根拆解或常見搭配詞（嚴格限制 20 字以內，不可換行）。禁止重複中文解釋。範例：'over(超過)+haul(拉)=徹底翻修' 或 '常搭配 undergo（經歷）'
- "sentence_en" : an English example sentence
- "sentence_ipa": full IPA pronunciation of the example sentence
- "sentence_cn" : Traditional Chinese translation of the example sentence"""

    if context:
        # ════════════════════════════════════════════════════
        # 🚀 引擎 A：知識與新聞模式 (Master Prompt)
        # ════════════════════════════════════════════════════
        print(f"🤖 正在根據來源生成「{topic}」的進階知識卡片...")

        # 防呆：> 100 字元才算有效文本，短內容改走腦力激盪
        has_valid_context = len(context) > 100

        prompt = f"""You are an expert English vocabulary curator for advanced learners.
Your mission: extract or generate {count} ADVANCED vocabulary items that would genuinely help learners sound fluent and professional in English media, tech writing, or business contexts.

Topic/Focus: "{topic}"
"""
        if has_valid_context:
            prompt += f"""
Source Material:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{context}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
STRICT RULE: Every vocabulary item and example sentence MUST be grounded in the source material above. Use actual names, events, or concepts that appear in the source.
"""
        else:
            prompt += f"""
STRICT RULE: No specific source text is provided. You MUST brainstorm highly relevant, industry-specific, or advanced expressions related to the topic "{topic}". The example sentences MUST sound like they are from a professional news report, tech blog, or business analysis.
"""
        prompt += f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
STRICT SELECTION RULES:
✅ PRIORITIZE: Multi-word collocations, industry jargon, and precise single words.
❌ FORBIDDEN: Basic vocabulary (launch, new, use, make, show, say, oops, cringe, bad, sad), generic filler phrases.

SENTENCE RULE: Each "sentence_en" MUST contain a concrete detail and sound like professional journalism.

CONTEXT FIT RULE: Every word/phrase MUST be naturally and logically appropriate for the topic "{topic}". If a word feels forced or unnatural in this context, either adjust the example sentence to a scenario where the word fits naturally, or replace the word with a more contextually fitting alternative. NEVER force an example sentence to justify an out-of-context word.

TRANSLATION RULE: All Chinese content (word_cn, sentence_cn) MUST use natural, colloquial Taiwanese Mandarin (台灣日常口語意譯). NEVER produce word-for-word literal translations from English. Use expressions a Taiwanese person would naturally say.

FREQUENCY RULE: Prioritize vocabulary that is GENUINELY HIGH-FREQUENCY in this field — words that appear regularly in industry publications, expert conversations, or job listings. Avoid impressive-sounding jargon that professionals rarely encounter in practice.
"""

    else:
        # ════════════════════════════════════════════════════
        # ☕ 引擎 B：日常與生活模式 (Everyday Prompt)
        # ════════════════════════════════════════════════════
        print(f"💬 正在生成「{topic}」的日常實用生活卡片...")

        prompt = f"""You are an enthusiastic and practical English vocabulary teacher.
Your mission: Generate {count} highly useful, everyday English vocabulary items (words, phrases, or idioms) related to the topic "{topic}".

VARIETY RULE: Spread across different real-life scenarios — home, work, socializing, shopping, health, emotions, dining, transport, requests, small talk. Do NOT cluster around one scenario.

STRICT SELECTION RULES:
✅ PRIORITIZE: Natural spoken expressions, phrasal verbs, idioms, and collocations that native speakers use every day in real conversations.
❌ FORBIDDEN:
  1. Overly academic, technical, or travel-specific vocabulary
  2. Stiff, textbook-style words or sentences
  3. Any word or phrase you used as an example in previous responses — always pick fresh vocabulary

SENTENCE RULE: Each "sentence_en" MUST be a practical, real-life sentence that a beginner can immediately use in daily conversation — something people actually say at work, at a restaurant, with friends, or while traveling. Prioritize usefulness over formality.

CONTEXT FIT RULE: Every word/phrase MUST be naturally and logically appropriate for the topic "{topic}". If a word feels forced or unnatural in this context, either adjust the example sentence to a scenario where the word fits naturally, or replace the word with a more contextually fitting alternative. NEVER force an example sentence to justify an out-of-context word.

TRANSLATION RULE: All Chinese content (word_cn, sentence_cn) MUST use natural, colloquial Taiwanese Mandarin (台灣日常口語意譯). NEVER produce word-for-word literal translations from English. Use expressions a Taiwanese person would naturally say.

FREQUENCY RULE: Prioritize vocabulary that is HIGH-FREQUENCY in real daily life. Every word/phrase should be something a learner will realistically hear or use within their first year of speaking English. Avoid rare or obscure expressions.
"""

    # ── 跨次歷史紀錄 ──────────────────────────────────────
    used_words: set[str] = set()
    if os.path.exists(USED_WORDS_FILE):
        try:
            with open(USED_WORDS_FILE, "r", encoding="utf-8") as f:
                used_words = set(json.load(f))
        except Exception:
            pass

    def _call_chunk(chunk_size: int, exclude: set[str]) -> list:
        """呼叫一次 OpenAI，回傳去重後的新詞彙。"""
        exclusion_note = ""
        if exclude:
            # 隨機抽樣，避免永遠只排除 Z 開頭的詞
            sample = random.sample(list(exclude), min(200, len(exclude)))
            exclusion_note = f"\nNEVER generate any of these already-used words/phrases: {', '.join(sample)}\n"

        full_prompt = prompt + exclusion_note + f"\nGenerate exactly {chunk_size} items.\n" + field_spec
        try:
            resp = _call_openai(
                messages=[{"role": "user", "content": full_prompt}],
                model="gpt-4o-mini",
                response_format={"type": "json_object"},
                temperature=1.1,
            )
            raw = json.loads(resp.choices[0].message.content).get("items", [])
        except Exception as e:
            print(f"      [Chunk 警告] API 呼叫或解析失敗，等待 2 秒後重試: {e}")
            time.sleep(2)
            return []

        result = []
        for item in raw:
            # 標點符號正規化："word." 和 "word" 視為同一詞
            key = item.get("word_en", "").lower().strip().strip(".,!?")
            if key and key not in exclude:
                if item.get("word_en"):
                    w = item["word_en"]
                    item["word_en"] = w[0].upper() + w[1:] if w else w
                result.append(item)
                exclude.add(key)
        return result

    # ── 分批生成直到數量足夠 ──────────────────────────────
    all_items: list   = []
    seen: set[str]    = set(used_words)
    max_rounds        = math.ceil(count / GENERATE_CHUNK) + 3
    rounds            = 0
    consecutive_fails = 0

    while len(all_items) < count and rounds < max_rounds:
        need       = count - len(all_items)
        chunk_size = min(GENERATE_CHUNK, need + 3)
        print(f"   🔄 OpenAI 呼叫 #{rounds + 1}（目標 {chunk_size} 個，已有 {len(all_items)} 個）...")
        new_items = _call_chunk(chunk_size, seen)
        if not new_items:
            consecutive_fails += 1
            if consecutive_fails >= 3:
                print("   ⚠️  連續失敗 3 次，中斷生成。")
                break
            continue   # API 失敗不消耗 rounds
        consecutive_fails = 0
        all_items.extend(new_items)
        rounds += 1

    all_items = all_items[:count]

    for i, item in enumerate(all_items):
        item["id"] = f"{i + 1:02d}"

    # ── 更新歷史紀錄（原子寫入）──────────────────────────
    new_keys = {item["word_en"].lower().strip().strip(".,!?") for item in all_items if item.get("word_en")}
    updated  = used_words | new_keys
    try:
        tmp_words_file = USED_WORDS_FILE + ".tmp"
        with open(tmp_words_file, "w", encoding="utf-8") as f:
            json.dump(sorted(updated), f, ensure_ascii=False, indent=2)
        os.replace(tmp_words_file, USED_WORDS_FILE)
    except Exception as e:
        print(f"   ⚠️  歷史紀錄儲存失敗: {e}")

    return all_items


def get_tavily_context(query: str, max_chars: int = 6000) -> str:
    """使用 Tavily API 搜尋主題背景知識，回傳整理好的文本供 LLM 參考。
    支援雙金鑰自動輪換。需要 .env 設定 TAVILY_API_KEY（與 TAVILY_API_KEY_2）。
    """
    if not TAVILY_KEYS:
        print("⚠️  未設定任何 TAVILY_API_KEY，將略過知識檢索")
        return ""
    try:
        from tavily import TavilyClient  # noqa: F401
    except ImportError:
        print("⚠️  tavily-python 未安裝，請執行: pip install tavily-python")
        return ""

    print(f"🔍 正在透過 Tavily 搜尋「{query}」相關知識...")
    try:
        context, key_num = _call_tavily(query, search_depth="advanced", max_tokens=max_chars)
        suffix = f"（備用金鑰 #{key_num}）" if key_num > 1 else ""
        print(f"   ✅ 已取得背景知識（{len(context)} 字元）{suffix}")
        return context
    except Exception as e:
        print(f"   ⚠️  所有 Tavily 金鑰均失敗，略過知識檢索")
        return ""


def _extract_youtube_id(url: str) -> str | None:
    """從 YouTube 完整 URL 或短網址中擷取 11 碼 video ID"""
    m = re.search(r'(?:youtube\.com/watch\?v=|youtu\.be/)([a-zA-Z0-9_-]{11})', url)
    return m.group(1) if m else None


def _get_youtube_transcript(url: str, max_chars: int = 15000) -> str:
    """
    使用 youtube-transcript-api 抓取字幕。
    先用 list() 列出所有可用軌道，再依優先順序挑選：
      1. 手動建立的英文字幕（en / en-US / en-GB …）
      2. 自動產生的英文字幕（同上語言變體）
      3. 任何第一條可用字幕
    回傳前 max_chars 字元的逐字稿。
    """
    try:
        from youtube_transcript_api import YouTubeTranscriptApi
    except ImportError:
        print("⚠️  youtube-transcript-api 未安裝，請執行: pip install youtube-transcript-api")
        return ""

    video_id = _extract_youtube_id(url)
    if not video_id:
        print("⚠️  無法解析 YouTube video ID")
        return ""

    print(f"🎬 正在擷取 YouTube 字幕（video: {video_id}）...")
    try:
        transcripts = list(YouTubeTranscriptApi().list(video_id))
    except Exception as e:
        print(f"   ⚠️  無法取得字幕列表: {e}")
        return ""

    en_variants = {"en", "en-US", "en-GB", "en-CA", "en-AU"}

    # 1. 手動英文
    transcript = next((t for t in transcripts if t.language_code in en_variants and not t.is_generated), None)

    # 2. 自動英文
    if transcript is None:
        transcript = next((t for t in transcripts if t.language_code in en_variants and t.is_generated), None)

    # 3. 任何第一條可用字幕
    if transcript is None:
        transcript = transcripts[0] if transcripts else None

    if transcript is None:
        print("   ⚠️  找不到可用字幕（影片可能未開放或無字幕）")
        return ""

    try:
        entries  = transcript.fetch()
        raw_text = " ".join(
            e.text if hasattr(e, "text") else e.get("text", "")
            for e in entries
        )
        result   = raw_text[:max_chars]
        kind     = "自動" if transcript.is_generated else "手動"
        print(f"   ✅ 字幕語言: {transcript.language_code}（{kind}），共 {len(result)} 字元")
        return result
    except Exception as e:
        print(f"   ⚠️  字幕擷取失敗: {e}")
        return ""


def _scrape_url(url: str, max_chars: int = 6000) -> str:
    """
    使用 requests + BeautifulSoup 抓取一般網頁的純文字內容。
    需要: pip install beautifulsoup4
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        print("⚠️  beautifulsoup4 未安裝，請執行: pip install beautifulsoup4")
        return ""

    print(f"🔗 正在抓取網頁內容...")
    try:
        resp = requests.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()
        text = re.sub(r'\s+', ' ', soup.get_text(separator=" ", strip=True))
        context = text[:max_chars]
        print(f"   ✅ 已抓取網頁內容（{len(context)} 字元）")
        return context
    except Exception as e:
        print(f"   ⚠️  網頁抓取失敗: {e}")
        return ""


def get_knowledge_context(query: str) -> str:
    """
    雙引擎智慧路由：
      引擎 A — YouTube URL → youtube-transcript-api 抓取字幕（最多 15000 字元）
      引擎 B — 一般關鍵字 → Tavily 網路搜尋（最多 6000 字元）
    若 A 失敗自動 fallback 至 B。
    """
    if re.search(r'youtube\.com|youtu\.be', query, re.IGNORECASE):
        context = _get_youtube_transcript(query)
        if context:
            return context
        print("   字幕擷取失敗，自動切換至 Tavily 搜尋引擎...")
    return get_tavily_context(query)


_SEARCH_PREFIXES = ("/search ", "/s ")

def route_input(user_input: str) -> tuple[str, str]:
    """
    五種輸入自動分流，回傳 (context, source_type)。

    source_type 對應：
      "youtube" — YouTube 字幕
      "url"     — 一般網頁爬取
      "text"    — 直接貼入長文
      "tavily"  — /search 或 /s 前綴觸發 Tavily 搜尋
      "normal"  — 一般主題，context = ""，LLM 走日常 prompt

    Firestore 分類建議：normal → "videofactory"，其餘 → "trending"
    """
    t = user_input.strip()

    # ── 1. YouTube URL ───────────────────────────────────
    if t.startswith("http") and re.search(r'youtube\.com|youtu\.be', t, re.IGNORECASE):
        print("🎬 偵測到 YouTube 網址，正在擷取字幕...")
        context = _get_youtube_transcript(t)
        if not context:
            print("   字幕擷取失敗，嘗試改用 Tavily 搜尋...")
            context = get_tavily_context(t)
        return context, "youtube"

    # ── 2. 一般網頁 URL ──────────────────────────────────
    if t.startswith("http"):
        print("🔗 偵測到一般網址，正在抓取網頁內容...")
        return _scrape_url(t), "url"

    # ── 3. 直接貼入長文（> 100 字元，非網址）────────────
    if len(t) > 100:
        print("📄 偵測到長文貼入，直接作為知識來源（前 6000 字元）...")
        return t[:6000], "text"

    # ── 4. Tavily 搜尋（/search 或 /s 前綴）─────────────
    lower = t.lower()
    for prefix in _SEARCH_PREFIXES:
        if lower.startswith(prefix):
            query = t[len(prefix):].strip()
            print(f"🔍 偵測到搜尋指令，正在透過 Tavily 搜尋「{query}」...")
            return get_tavily_context(query), "tavily"

    # ── 5. 一般主題文字 ──────────────────────────────────
    print("💬 一般主題模式，LLM 將依主題直接生成日常詞彙...")
    return "", "normal"


def download_pexels_images(topic: str, count: int = 10) -> list:
    """從 Pexels API 下載高品質橫向圖片，儲存至 temp/images/"""
    print(f"📸 正在從 Pexels 下載「{topic}」圖片（目標 {count} 張）...")

    headers = {"Authorization": PEXELS_KEY}
    params  = {
        "query":       topic,
        "per_page":    min(count, 80),
        "orientation": "landscape",
        "size":        "large",
    }

    try:
        resp = requests.get(
            "https://api.pexels.com/v1/search",
            headers=headers, params=params, timeout=15
        )
        resp.raise_for_status()
        photos = resp.json().get("photos", [])
    except Exception as e:
        print(f"   ⚠️  Pexels API 請求失敗: {e}，將使用本地圖片")
        return []

    if not photos:
        print(f"   ⚠️  搜尋「{topic}」無結果，將使用本地圖片")
        return []

    image_paths = []
    for i, photo in enumerate(photos[:count]):
        url      = photo["src"].get("large2x", photo["src"]["large"])
        img_path = os.path.join(IMAGES_DIR, f"pexels_{i:02d}.jpg")
        try:
            r = requests.get(url, timeout=30)
            r.raise_for_status()
            with open(img_path, "wb") as f:
                f.write(r.content)
            image_paths.append(img_path)
            print(f"   ✅ [{i + 1}/{len(photos[:count])}] 已下載")
        except Exception as e:
            print(f"   ⚠️  圖片 {i + 1} 下載失敗: {e}")

    return image_paths


# ================= 素材檢查 =================

def pick_bgm() -> str | None:
    """
    從 assets/bgm/ 資料夾隨機挑一首 MP3。
    若資料夾不存在或為空，fallback 至 assets/bgm.mp3。
    都找不到回傳 None。
    """
    if os.path.isdir(BGM_DIR):
        files = [
            os.path.join(BGM_DIR, f)
            for f in os.listdir(BGM_DIR)
            if f.lower().endswith(".mp3")
        ]
        if files:
            chosen = random.choice(files)
            print(f"🎵 本次 BGM: {os.path.basename(chosen)}")
            return chosen
    if os.path.exists(BGM_SINGLE):
        return BGM_SINGLE
    return None


def check_assets() -> bool:
    """確認必要素材存在（字型 + 至少一個 BGM 來源）"""
    missing = []
    if not os.path.exists(FONT_EN): missing.append("assets/font_en.ttf")
    if not os.path.exists(FONT_CN): missing.append("assets/font_cn.otf")

    has_bgm_dir = os.path.isdir(BGM_DIR) and any(
        f.lower().endswith(".mp3") for f in os.listdir(BGM_DIR)
    ) if os.path.isdir(BGM_DIR) else False

    if not has_bgm_dir and not os.path.exists(BGM_SINGLE):
        missing.append("assets/bgm/ 資料夾（執行 python download_bgm.py 下載）")

    if missing:
        print(f"❌ 缺少必要素材: {', '.join(missing)}")
        return False
    return True


def _get_fallback_bg_images() -> list:
    """取得本地備用背景圖（assets/bg/ 或 assets/bg.jpg）"""
    bg_dir = os.path.join(ASSETS_DIR, "bg")
    if os.path.isdir(bg_dir):
        imgs = [
            os.path.join(bg_dir, f)
            for f in os.listdir(bg_dir)
            if f.lower().endswith((".jpg", ".jpeg", ".png"))
        ]
        if imgs:
            return imgs
    fallback = os.path.join(ASSETS_DIR, "bg.jpg")
    return [fallback] if os.path.exists(fallback) else []


_last_bg_path: str | None = None   # 記錄上一張背景，避免前後連續相同


def _pick_bg(pexels_images: list) -> str:
    """從 Pexels 圖庫（優先）或本地備用隨機選一張背景，且前後不重複。"""
    global _last_bg_path
    pool = pexels_images if pexels_images else _get_fallback_bg_images()
    if not pool:
        raise FileNotFoundError(
            "找不到任何背景圖片！請確認 assets/bg.jpg 存在，或設定 PEXELS_API_KEY。"
        )
    candidates = [p for p in pool if p != _last_bg_path] if len(pool) > 1 else pool
    chosen = random.choice(candidates)
    _last_bg_path = chosen
    return chosen


# ================= 中文字型自動偵測 =================

def _best_cn_font_path() -> tuple[str, dict]:
    """
    偵測 font_cn.otf 的繁體中文字符覆蓋率。
    使用幾個常見缺字（繁體罕用字）做快速測試：
      · 覆蓋完整 → 繼續使用 assets/font_cn.otf
      · 出現缺字 → 自動切換至系統字型（PingFang / MS JhengHei）

    系統字型候選（macOS / Windows / Linux）均涵蓋所有常用繁體中文。
    """
    _SYSTEM_CANDIDATES = [
        ("/System/Library/Fonts/PingFang.ttc",                          {"index": 0}),  # macOS
        ("C:/Windows/Fonts/msjh.ttc",                                   {"index": 0}),  # Windows
        ("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",      {"index": 0}),  # Linux
        ("/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",      {"index": 0}),  # Linux alt
    ]
    # 測試字符：覆蓋率不足的字型通常在這些字上出現白匡
    _TEST_CHARS = "臺鬱纜鑑齡繽釀"

    if os.path.exists(FONT_CN):
        try:
            probe = ImageFont.truetype(FONT_CN, 20)
            if all(probe.getmask(ch).getbbox() is not None for ch in _TEST_CHARS):
                return FONT_CN, {}          # 覆蓋完整，使用自訂字型
            print("⚠️  font_cn.otf 字符集不完整（部分繁體字缺字），自動切換系統字型...")
        except Exception:
            pass

    for path, kwargs in _SYSTEM_CANDIDATES:
        if os.path.exists(path):
            print(f"   ✅ 使用系統字型: {os.path.basename(path)}")
            return path, kwargs

    print("⚠️  找不到完整中文字型，建議下載 Noto Sans CJK TC 放至 assets/font_cn.otf")
    return FONT_CN, {}


_CN_FONT_PATH, _CN_FONT_KWARGS = _best_cn_font_path()
_CN_FONT_CACHE: dict[int, ImageFont.FreeTypeFont] = {}


def _load_cn_font(size: int) -> ImageFont.FreeTypeFont:
    """中文字型快取載入（自動使用最佳字型，避免白匡）"""
    if size not in _CN_FONT_CACHE:
        _CN_FONT_CACHE[size] = ImageFont.truetype(_CN_FONT_PATH, size, **_CN_FONT_KWARGS)
    return _CN_FONT_CACHE[size]


# ================= 圖片繪製 =================

def _create_base_image(pexels_images: list) -> Image.Image:
    """建立 1920×1080 帶有 80% 暗色遮罩的背景圖"""
    bg_path = _pick_bg(pexels_images)
    base    = Image.open(bg_path).convert("RGBA").resize((1920, 1080))
    overlay = Image.new("RGBA", base.size, (0, 0, 0, 200))   # alpha=200 ≈ 78%
    return Image.alpha_composite(base, overlay)


def draw_text_wrapped(draw, text, font, max_width, start_x, start_y, color, line_spacing=15):
    """
    自動換行繪圖（支援中英文）。
    回傳：繪製結束後的 Y 座標（供後續文字接續使用）。
    """
    lines = []
    if any("\u4e00" <= c <= "\u9fff" for c in text):
        # 中文：以字元數截斷
        lines = textwrap.wrap(text, width=22)
    else:
        # 英文：以像素寬度截斷
        words = text.split()
        current_line: list[str] = []
        for word in words:
            test = " ".join(current_line + [word])
            w = draw.textbbox((0, 0), test, font=font)[2]
            if w <= max_width:
                current_line.append(word)
            else:
                if current_line:
                    lines.append(" ".join(current_line))
                current_line = [word]
        if current_line:
            lines.append(" ".join(current_line))

    current_y = start_y
    for line in lines:
        draw.text((start_x, current_y), line, font=font, fill=color)
        bbox = draw.textbbox((0, 0), line, font=font)
        current_y += (bbox[3] - bbox[1]) + line_spacing
    return current_y


def _normalize_token(w: str) -> str:
    """移除所有非英文字母，轉小寫（用於高亮匹配）"""
    return re.sub(r'[^a-z]', '', w.lower())


def _stem_token(tok: str) -> str:
    """簡易詞幹提取：移除常見英文字尾（用於 fallback 高亮匹配）"""
    for suffix in ("ings", "ing", "tion", "ness", "ment", "er", "est", "ed", "es", "s"):
        if tok.endswith(suffix) and len(tok) - len(suffix) >= 3:
            return tok[:-len(suffix)]
    return tok


def draw_text_with_highlight(
    draw, text, highlight_word, font, max_width,
    start_x, start_y, default_color, highlight_color, line_spacing=15
):
    """
    自動換行並對指定片語標記顏色（連續視窗精確匹配 + fallback 逐詞匹配）。
    回傳：繪製結束後的 Y 座標。
    """
    words    = text.split()
    s_tokens = [_normalize_token(w) for w in words]

    # 1. 連續視窗精確匹配
    h_tokens = [_normalize_token(w) for w in highlight_word.split() if _normalize_token(w)]
    highlighted: set[int] = set()
    n = len(h_tokens)
    for i in range(len(s_tokens) - n + 1):
        if s_tokens[i:i + n] == h_tokens:
            highlighted.update(range(i, i + n))

    # 2. fallback：完全沒匹配到時，改用詞幹匹配（應對 GPT 造句用了變形詞）
    if not highlighted:
        h_stems = {_stem_token(t) for t in h_tokens}
        for i, tok in enumerate(s_tokens):
            if _stem_token(tok) in h_stems and len(tok) > 2:
                highlighted.add(i)

    current_x, current_y, last_h = start_x, start_y, 0
    for idx, word in enumerate(words):
        color = highlight_color if idx in highlighted else default_color
        bbox  = draw.textbbox((0, 0), word + " ", font=font)
        w, h  = bbox[2] - bbox[0], bbox[3] - bbox[1]
        last_h = h
        if current_x + w > start_x + max_width:
            current_x  = start_x
            current_y += h + line_spacing
        draw.text((current_x, current_y), word + " ", font=font, fill=color)
        current_x += w

    return current_y + last_h + line_spacing


def create_word_card_image(data: dict, output_filename: str, pexels_images: list):
    """生成單字教學卡片（詞彙 + IPA + 中文 + Tip）"""
    base = _create_base_image(pexels_images)
    draw = ImageDraw.Draw(base)

    font_header = ImageFont.truetype(FONT_EN, 40)
    font_en     = ImageFont.truetype(FONT_EN, 100)
    font_ipa    = ImageFont.truetype(FONT_EN, 50)
    font_cn     = _load_cn_font(70)
    font_tips   = _load_cn_font(50)

    draw.text((100, 80), f"{data['id']}  |  {BRAND_NAME} - Vocabulary", font=font_header, fill="white")

    sx, mw, cy = 150, 1600, 250
    cy = draw_text_wrapped(draw, data["word_en"],  font_en,   mw, sx, cy, "#ffdd00", 20)
    cy += 20
    cy = draw_text_wrapped(draw, data["word_ipa"], font_ipa,  mw, sx, cy, "#cccccc", 15)
    cy += 60
    cy = draw_text_wrapped(draw, data["word_cn"],  font_cn,   mw, sx, cy, "white",   20)
    cy += 100
    draw.text((sx, cy), "Tip:", font=font_tips, fill="#00ffcc")
    cy += 60
    draw_text_wrapped(draw, data["tips"], font_tips, mw, sx, cy, "#eeeeee", 20)

    base.save(output_filename)


def create_sentence_card_image(data: dict, output_filename: str, pexels_images: list):
    """生成例句教學卡片（例句高亮 + 中文翻譯）"""
    base = _create_base_image(pexels_images)
    draw = ImageDraw.Draw(base)

    font_header = ImageFont.truetype(FONT_EN, 40)
    font_en     = ImageFont.truetype(FONT_EN, 80)
    font_cn     = _load_cn_font(70)

    draw.text((100, 80), f"{data['id']}  |  {BRAND_NAME} - Example", font=font_header, fill="white")

    sx, mw, cy = 150, 1600, 300
    cy = draw_text_with_highlight(
        draw, data["sentence_en"], data["word_en"],
        font=font_en, max_width=mw, start_x=sx, start_y=cy,
        default_color="white", highlight_color="#ffdd00", line_spacing=20,
    )
    cy += 40
    draw_text_wrapped(draw, data["sentence_cn"], font_cn, mw, sx, cy, "white", 20)
    base.save(output_filename)


def _apply_frosted_glass(img: Image.Image, hidden_y: int):
    """對 img 的 hidden_y~1080 區域套用磨砂玻璃效果（模糊 + 半透明遮罩）"""
    region  = img.crop((0, hidden_y, 1920, 1080))
    blurred = region.filter(ImageFilter.GaussianBlur(radius=18))
    overlay = Image.new("RGBA", blurred.size, (0, 0, 0, 120))
    blurred = Image.alpha_composite(blurred, overlay)
    img.paste(blurred, (0, hidden_y))


def create_word_card_hidden_image(data: dict, output_filename: str, pexels_images: list):
    """生成單字主動回憶卡片（中文+tip 以磨砂玻璃遮住）"""
    base = _create_base_image(pexels_images)
    draw = ImageDraw.Draw(base)

    font_header = ImageFont.truetype(FONT_EN, 40)
    font_en     = ImageFont.truetype(FONT_EN, 100)
    font_ipa    = ImageFont.truetype(FONT_EN, 50)
    font_cn     = _load_cn_font(70)
    font_tips   = _load_cn_font(50)

    draw.text((100, 80), f"{data['id']}  |  {BRAND_NAME} - Active Recall", font=font_header, fill="white")

    sx, mw, cy = 150, 1600, 250
    cy = draw_text_wrapped(draw, data["word_en"],  font_en,  mw, sx, cy, "#ffdd00", 20)
    cy += 20
    cy = draw_text_wrapped(draw, data["word_ipa"], font_ipa, mw, sx, cy, "#cccccc", 15)
    cy += 60

    hidden_y = cy  # 中文從此處開始 → 模糊遮罩
    cy = draw_text_wrapped(draw, data["word_cn"], font_cn,   mw, sx, cy, "white",   20)
    cy += 100
    draw.text((sx, cy), "Tip:", font=font_tips, fill="#00ffcc")
    cy += 60
    draw_text_wrapped(draw, data["tips"], font_tips, mw, sx, cy, "#eeeeee", 20)

    _apply_frosted_glass(base, hidden_y)
    hint_draw = ImageDraw.Draw(base)
    font_hint  = _load_cn_font(55)
    hint_draw.text((150, hidden_y + 25), "請回想中文意思", font=font_hint, fill="#ffe066")
    base.save(output_filename)


def create_sent_card_hidden_image(data: dict, output_filename: str, pexels_images: list):
    """生成例句主動回憶卡片（中文以磨砂玻璃遮住）"""
    base = _create_base_image(pexels_images)
    draw = ImageDraw.Draw(base)

    font_header = ImageFont.truetype(FONT_EN, 40)
    font_en     = ImageFont.truetype(FONT_EN, 80)
    font_ipa    = ImageFont.truetype(FONT_EN, 50)
    font_cn     = _load_cn_font(70)

    draw.text((100, 80), f"{data['id']}  |  {BRAND_NAME} - Active Recall", font=font_header, fill="white")

    sx, mw, cy = 150, 1600, 300
    cy = draw_text_with_highlight(
        draw, data["sentence_en"], data["word_en"],
        font=font_en, max_width=mw, start_x=sx, start_y=cy,
        default_color="white", highlight_color="#ffdd00", line_spacing=20,
    )
    cy += 40

    hidden_y = cy  # 中文從此處開始 → 模糊遮罩
    draw_text_wrapped(draw, data["sentence_cn"], font_cn, mw, sx, cy, "white", 20)

    _apply_frosted_glass(base, hidden_y)
    hint_draw = ImageDraw.Draw(base)
    font_hint  = _load_cn_font(55)
    hint_draw.text((150, hidden_y + 25), "請回想中文意思", font=font_hint, fill="#ffe066")
    base.save(output_filename)


# ================= 音訊合成 =================

def clean_for_tts(text: str) -> str:
    """移除會被 TTS 誤讀的標點（e.g. '-' 被中文唸成「負」，'+' 唸成「加」或「plus」）"""
    text = re.sub(r"[\-/+]", " ", text)
    text = re.sub(r"[()（）]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


async def generate_audio(text: str, voice: str, output_filename: str, retries: int = 3, rate: str = "+0%"):
    """使用 Edge-TTS 生成語音，含自動重試機制。rate: '+0%'=原速, '-20%'=0.8x慢速"""
    for attempt in range(retries):
        try:
            comm = edge_tts.Communicate(text, voice, rate=rate)
            await comm.save(output_filename)
            return
        except Exception as e:
            if attempt == retries - 1:
                raise
            print(f"      [警告] TTS 失敗 ({e})，第 {attempt + 1} 次重試...")
            await asyncio.sleep(2)


# ================= 時間工具 =================

def _srt_time(seconds: float) -> str:
    """秒數 → SRT 時間格式 HH:MM:SS,mmm"""
    ms = int(round((seconds % 1) * 1000))
    s  = int(seconds) % 60
    m  = int(seconds) // 60 % 60
    h  = int(seconds) // 3600
    return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"


def _chapter_time(seconds: float) -> str:
    """秒數 → YouTube 章節格式 MM:SS"""
    m = int(seconds) // 60
    s = int(seconds) % 60
    return f"{m:02d}:{s:02d}"


def _normalize_prebuilt(src: str) -> str:
    """
    將 pre-built 影片轉碼為與 chunk 相同規格：1920x1080, 24fps, libx264, aac 48kHz。
    輸出至 TEMP_DIR，只在規格不符時才轉碼（用 ffprobe 檢查）。
    回傳正規化後的路徑。
    """
    try:
        result = subprocess.run(
            ["ffprobe", "-v", "quiet", "-print_format", "json",
             "-show_streams", src],
            capture_output=True, text=True, check=True,
        )
        streams = json.loads(result.stdout).get("streams", [])
    except Exception:
        streams = []

    needs_transcode = False
    for s in streams:
        if s.get("codec_type") == "video":
            w = int(s.get("width", 0))
            h = int(s.get("height", 0))
            # fps 可能是 "24/1" 或 "24000/1001"
            r_str = s.get("r_frame_rate", "0/1")
            try:
                num, den = r_str.split("/")
                fps_val = float(num) / float(den)
            except Exception:
                fps_val = 0.0
            if w != 1920 or h != 1080 or abs(fps_val - 24) > 0.5:
                needs_transcode = True
        if s.get("codec_type") == "audio":
            if int(s.get("sample_rate", 0)) != 48000:
                needs_transcode = True

    if not needs_transcode:
        return src

    basename = os.path.basename(src)
    dst = os.path.join(TEMP_DIR, f"normalized_{basename}")
    if os.path.exists(dst):
        print(f"♻️  已存在正規化版本：{basename}")
        return dst

    print(f"🔄 正在正規化 pre-built 影片規格：{basename} ...")
    ret = os.system(
        f'ffmpeg -y -i "{src}" '
        f'-vf "scale=1920:1080:force_original_aspect_ratio=decrease,'
        f'pad=1920:1080:(ow-iw)/2:(oh-ih)/2" '
        f'-r 24 -c:v libx264 -preset fast '
        f'-c:a aac -ar 48000 -ac 2 -b:a 192k '
        f'"{dst}" -loglevel error'
    )
    if ret != 0 or not os.path.exists(dst):
        print(f"⚠️  正規化失敗，改用原始檔案：{basename}")
        return src

    print(f"✅ 已正規化：{basename}")
    return dst


def _video_duration(path: str) -> float:
    """用 ffprobe 快速取得影片時長（秒）"""
    try:
        result = subprocess.run(
            ["ffprobe", "-v", "quiet", "-print_format", "json", "-show_format", path],
            capture_output=True, text=True, check=True,
        )
        info = json.loads(result.stdout)
        return float(info["format"]["duration"])
    except Exception:
        # 備用：用 MoviePy 讀取
        clip = VideoFileClip(path)
        dur  = clip.duration
        clip.close()
        return dur


# ================= 批次處理 =================

async def process_group(
    group:          list,
    group_idx:      int,
    pexels_images:  list,
    cumulative_time: float,
    srt_entries:    list,
    chapter_entries: list,
    phase:          int = 1,
) -> tuple[list, float]:
    """
    處理一組教學項目：生成圖片 → 生成音訊 → 組合片段 → 匯出 Batch MP4。
    回傳: (chunk_file_list, updated_cumulative_time)
    """
    chunk_files = []

    # ── Checkpoint 恢復 ───────────────────────────────────
    ckpt_file = os.path.join(TEMP_DIR, f"ckpt_g{group_idx}.json")
    ckpt: dict = {}
    if os.path.exists(ckpt_file):
        try:
            with open(ckpt_file, "r", encoding="utf-8") as _f:
                ckpt = json.load(_f)
            print(f"   ♻️  偵測到 checkpoint（已完成 {len(ckpt)} 個 batch）")
        except Exception:
            ckpt = {}

    for batch_num, batch_start in enumerate(range(0, len(group), BATCH_SIZE)):
        batch      = group[batch_start:batch_start + BATCH_SIZE]
        batch_label = f"G{group_idx}-B{batch_num}"
        chunk_path  = os.path.join(TEMP_DIR, f"chunk_g{group_idx}_b{batch_num}.mp4")

        # ── 已有 checkpoint，直接恢復 ──────────────────────
        key = str(batch_num)
        if key in ckpt and os.path.exists(chunk_path):
            entry = ckpt[key]
            cumulative_time = entry["cumulative_time_after"]
            srt_entries.extend(entry["srt_entries_delta"])
            chapter_entries.extend(entry["chapter_entries_delta"])
            chunk_files.append(chunk_path)
            print(f"\n⏩ Batch {batch_label} 已有 checkpoint，跳過")
            continue

        print(f"\n📦 Batch {batch_label}（項目 {batch_start + 1}～{batch_start + len(batch)}）")

        srt_before     = len(srt_entries)
        chapter_before = len(chapter_entries)
        batch_clips    = []

        for i, item in enumerate(batch):
            global_idx = group_idx * 1000 + batch_start + i   # unique temp-file prefix
            print(f"   [{item['id']}] {item['word_en']}")

            # 章節標記（以此詞彙開始的時間點）
            phase_label = "" if phase == 1 else "🔄 "
            chapter_entries.append((cumulative_time, f"{phase_label}{item['id']} - {item['word_en']}"))

            # ── 檔案路徑 ──────────────────────────────────
            img_word    = os.path.join(TEMP_DIR, f"card_word_{global_idx}.png")
            img_sent    = os.path.join(TEMP_DIR, f"card_sent_{global_idx}.png")
            aud_word_en = os.path.join(TEMP_DIR, f"word_en_{global_idx}.mp3")
            aud_word_cn = os.path.join(TEMP_DIR, f"word_cn_{global_idx}.mp3")
            aud_sent_en = os.path.join(TEMP_DIR, f"sent_en_{global_idx}.mp3")
            aud_sent_cn = os.path.join(TEMP_DIR, f"sent_cn_{global_idx}.mp3")

            try:
                if phase == 1:
                    # ── Phase 1：完整教學 ─────────────────────────
                    aud_tips         = os.path.join(TEMP_DIR, f"tips_{global_idx}.mp3")
                    aud_sent_en_slow = os.path.join(TEMP_DIR, f"sent_en_slow_{global_idx}.mp3")

                    # A. 生成圖片
                    create_word_card_image(item, img_word, pexels_images)
                    create_sentence_card_image(item, img_sent, pexels_images)

                    # B. 生成音訊
                    await asyncio.gather(
                        generate_audio(clean_for_tts(item["word_en"]),     "en-US-GuyNeural",       aud_word_en,      rate="-10%"),
                        generate_audio(clean_for_tts(item["word_cn"]),     "zh-TW-HsiaoChenNeural", aud_word_cn),
                        generate_audio(clean_for_tts(item["tips"]),        "zh-TW-HsiaoChenNeural", aud_tips),
                        generate_audio(clean_for_tts(item["sentence_en"]), "en-US-GuyNeural",       aud_sent_en_slow, rate="-20%"),
                        generate_audio(clean_for_tts(item["sentence_en"]), "en-US-GuyNeural",       aud_sent_en,      rate="-10%"),
                        generate_audio(clean_for_tts(item["sentence_cn"]), "zh-TW-HsiaoChenNeural", aud_sent_cn),
                    )

                    # C. 讀取 AudioClip（用 _load_audio 取得 ffprobe 精確時長）
                    c_w_en      = _load_audio(aud_word_en)
                    c_w_cn      = _load_audio(aud_word_cn)
                    c_tips      = _load_audio(aud_tips)
                    c_s_en_slow = _load_audio(aud_sent_en_slow)
                    c_s_en      = _load_audio(aud_sent_en)
                    c_s_cn      = _load_audio(aud_sent_cn)
                    c_s_cn2     = _load_audio(aud_sent_cn)   # 第二輪專用，避免共用 ffmpeg pipe

                    dur_w_en      = c_w_en.duration
                    dur_w_cn      = c_w_cn.duration
                    dur_tips      = c_tips.duration
                    dur_s_en_slow = c_s_en_slow.duration
                    dur_s_en      = c_s_en.duration
                    dur_s_cn      = c_s_cn.duration

                    # D. 記錄 SRT 時間戳（以聲音為主，兩次校準點）
                    # 校準點 1：單字段結束（word_en + word_cn + tips 音訊精確總和）
                    word_section = dur_w_en + dur_w_cn + dur_tips
                    sent_start   = cumulative_time + word_section
                    # 校準點 2：例句段結束（sent 音訊精確總和）
                    round1       = dur_s_en_slow + dur_s_cn
                    round2       = dur_s_en      + dur_s_cn
                    sent_end     = sent_start + round1 + round2

                    srt_entries.append((sent_start, sent_end,
                                        item["sentence_en"], item["sentence_cn"]))
                    cumulative_time += word_section + round1 + round2

                    # E. 組合影片片段（單字卡 → 慢速EN→CN → 原速EN→CN）
                    v_w_en  = _image_clip(img_word, c_w_en)
                    v_w_cn  = _image_clip(img_word, c_w_cn)
                    v_tips  = _image_clip(img_word, c_tips)

                    # ── karaoke：英文例句用「有CN無EN底圖」，karaoke 在 y=820 疊上英文 ──
                    kar_slow_path = os.path.join(TEMP_DIR, f"kar_sent_en_slow_{global_idx}.mp4")
                    kar_norm_path = os.path.join(TEMP_DIR, f"kar_sent_en_{global_idx}.mp4")
                    kar_slow = create_rounded_highlight_video(
                        img_sent, aud_sent_en_slow, kar_slow_path, FONT_EN, item["word_en"]
                    )
                    kar_norm = create_rounded_highlight_video(
                        img_sent, aud_sent_en, kar_norm_path, FONT_EN, item["word_en"]
                    )
                    # karaoke 成功 → 用 VideoFileClip；失敗 → 降級回靜態卡片
                    v_s_en_slow = VideoFileClip(kar_slow_path) if kar_slow else _image_clip(img_sent, c_s_en_slow)
                    v_s_en      = VideoFileClip(kar_norm_path) if kar_norm else _image_clip(img_sent, c_s_en)
                    v_s_cn      = _image_clip(img_sent, c_s_cn)
                    v_s_cn2     = _image_clip(img_sent, c_s_cn2)

                    item_clip = concatenate_videoclips([
                        v_w_en, v_w_cn, v_tips,
                        v_s_en_slow, v_s_cn,
                        v_s_en,      v_s_cn2,
                    ])

                else:
                    # ── Phase 2：Active Recall ────────────────────
                    img_word_hidden   = os.path.join(TEMP_DIR, f"card_word_hidden_{global_idx}.png")
                    img_sent_hidden   = os.path.join(TEMP_DIR, f"card_sent_hidden_{global_idx}.png")
                    aud_word_en_slow  = os.path.join(TEMP_DIR, f"word_en_slow_{global_idx}.mp3")
                    aud_sent_en_slow  = os.path.join(TEMP_DIR, f"sent_en_slow2_{global_idx}.mp3")
                    aud_sent_en_norm  = os.path.join(TEMP_DIR, f"sent_en_norm_{global_idx}.mp3")

                    # A. 生成圖片（完整 + 遮中文）
                    create_word_card_image(item, img_word, pexels_images)
                    create_word_card_hidden_image(item, img_word_hidden, pexels_images)
                    create_sentence_card_image(item, img_sent, pexels_images)
                    create_sent_card_hidden_image(item, img_sent_hidden, pexels_images)

                    # B. 生成音訊
                    # 步驟1：0.8x word_en（遮中文，思考用）
                    # 步驟2：word_cn（揭曉）
                    # 步驟3：0.8x sent_en（遮中文，思考用）
                    # 步驟4：sent_cn（揭曉）
                    # 步驟5：原速 sent_en（已揭曉，覆誦）
                    # 步驟6：sent_cn（鞏固收尾）
                    await asyncio.gather(
                        generate_audio(clean_for_tts(item["word_en"]),     "en-US-GuyNeural",       aud_word_en_slow, rate="-20%"),
                        generate_audio(clean_for_tts(item["word_cn"]),     "zh-TW-HsiaoChenNeural", aud_word_cn),
                        generate_audio(clean_for_tts(item["sentence_en"]), "en-US-GuyNeural",       aud_sent_en_slow, rate="-20%"),
                        generate_audio(clean_for_tts(item["sentence_en"]), "en-US-GuyNeural",       aud_sent_en_norm, rate="-10%"),
                        generate_audio(clean_for_tts(item["sentence_cn"]), "zh-TW-HsiaoChenNeural", aud_sent_cn),
                    )

                    # C. 讀取 AudioClip（sent_cn 需兩份：步驟4 & 步驟6；用 _load_audio 取精確時長）
                    c_w_en_slow  = _load_audio(aud_word_en_slow)
                    c_w_cn       = _load_audio(aud_word_cn)
                    c_s_en_slow  = _load_audio(aud_sent_en_slow)
                    c_s_en_norm  = _load_audio(aud_sent_en_norm)
                    c_s_cn       = _load_audio(aud_sent_cn)
                    c_s_cn2      = _load_audio(aud_sent_cn)

                    dur_w_en_slow  = c_w_en_slow.duration
                    dur_w_cn       = c_w_cn.duration
                    dur_s_en_slow  = c_s_en_slow.duration
                    dur_s_en_norm  = c_s_en_norm.duration
                    dur_s_cn       = c_s_cn.duration

                    # D. 記錄 SRT 時間戳（以聲音為主，RECALL_THINK_TIME 保留為有意義的思考留白）
                    word_section = (dur_w_en_slow + RECALL_THINK_TIME) + dur_w_cn
                    sent_start   = cumulative_time + word_section
                    round_recall = (dur_s_en_slow + RECALL_THINK_TIME) + dur_s_cn
                    round_norm   = dur_s_en_norm + c_s_cn2.duration
                    sent_end     = sent_start + round_recall + round_norm

                    srt_entries.append((sent_start, sent_end,
                                        item["sentence_en"], item["sentence_cn"]))
                    cumulative_time += word_section + round_recall + round_norm

                    # E. 組合影片片段（6 步 Active Recall）
                    # 1. 遮中文單字卡 + word_en (0.8x) + 2s靜音 → 聽英文，回想中文
                    # 2. 完整單字卡   + word_cn           → 揭曉解答
                    # 3. karaoke EN（無CN底圖）+ 2s靜態思考留白 → 聽例句，回想中文
                    # 4. 完整例句卡   + sent_cn           → 揭曉中文
                    # 5. karaoke EN（有CN底圖）→ 覆誦
                    # 6. 完整例句卡   + sent_cn           → 鞏固收尾
                    v1 = _image_clip(img_word_hidden, c_w_en_slow, extra_dur=RECALL_THINK_TIME)
                    v2 = _image_clip(img_word,        c_w_cn)

                    # v3：karaoke（無CN）+ 2s 靜態思考留白
                    kar2_slow_path = os.path.join(TEMP_DIR, f"kar2_sent_en_slow_{global_idx}.mp4")
                    kar2_slow = create_rounded_highlight_video(
                        img_sent_hidden, aud_sent_en_slow, kar2_slow_path, FONT_EN, item["word_en"]
                    )
                    if kar2_slow:
                        _fps_a   = 44100
                        _silence = AudioClip(lambda t: [0, 0], duration=RECALL_THINK_TIME, fps=_fps_a)
                        _think   = ImageClip(img_sent_hidden).with_duration(RECALL_THINK_TIME).with_audio(_silence)
                        v3 = concatenate_videoclips([VideoFileClip(kar2_slow_path), _think])
                    else:
                        v3 = _image_clip(img_sent_hidden, c_s_en_slow, extra_dur=RECALL_THINK_TIME)

                    v4 = _image_clip(img_sent, c_s_cn)

                    # v5：karaoke（有CN）→ 覆誦
                    kar2_norm_path = os.path.join(TEMP_DIR, f"kar2_sent_en_norm_{global_idx}.mp4")
                    kar2_norm = create_rounded_highlight_video(
                        img_sent, aud_sent_en_norm, kar2_norm_path, FONT_EN, item["word_en"]
                    )
                    v5 = VideoFileClip(kar2_norm_path) if kar2_norm else _image_clip(img_sent, c_s_en_norm)

                    v6 = _image_clip(img_sent, c_s_cn2)

                    item_clip = concatenate_videoclips([v1, v2, v3, v4, v5, v6])

                batch_clips.append(item_clip)

            except Exception as e:
                print(f"   ⚠️  項目 {item['id']} 合成失敗: {e}")

        # ── F. 匯出 Batch 影片並釋放記憶體 ─────────────
        if batch_clips:
            chunk_path = os.path.join(TEMP_DIR, f"chunk_g{group_idx}_b{batch_num}.mp4")
            print(f"   => 匯出 {os.path.basename(chunk_path)} ...")
            batch_video = concatenate_videoclips(batch_clips, method="chain")
            batch_video.write_videofile(
                chunk_path,
                fps=FPS,
                codec="libx264",
                audio_codec="aac",
                audio_fps=48000,
                threads=4,
                preset="ultrafast",
                logger=None,
            )
            chunk_files.append(chunk_path)

            # ── 儲存 Checkpoint ────────────────────────────
            ckpt[str(batch_num)] = {
                "cumulative_time_after":  cumulative_time,
                "srt_entries_delta":      srt_entries[srt_before:],
                "chapter_entries_delta":  chapter_entries[chapter_before:],
            }
            try:
                tmp_ckpt_file = ckpt_file + ".tmp"
                with open(tmp_ckpt_file, "w", encoding="utf-8") as _f:
                    json.dump(ckpt, _f, ensure_ascii=False, indent=2)
                os.replace(tmp_ckpt_file, ckpt_file)
            except Exception as _e:
                print(f"   ⚠️  checkpoint 儲存失敗: {_e}")

            batch_video.close()
            for clip in batch_clips:
                clip.close()
            gc.collect()

    # 清理 MP3 轉換產生的 WAV 暫存檔
    for ext in ("*.wav", "*.mp3", "*.png"):
        for f in glob.glob(os.path.join(TEMP_DIR, ext)):
            try:
                os.remove(f)
            except OSError:
                pass

    # 全部完成後移除 checkpoint（下次重新開始）
    if os.path.exists(ckpt_file):
        try:
            os.remove(ckpt_file)
        except OSError:
            pass

    return chunk_files, cumulative_time


# ================= 輸出檔案 =================

def write_srt(srt_entries: list, output_path: str):
    """將時間戳記錄寫出為標準 SRT 字幕檔"""
    with open(output_path, "w", encoding="utf-8") as f:
        for idx, (start, end, en_text, cn_text) in enumerate(srt_entries, 1):
            f.write(f"{idx}\n")
            f.write(f"{_srt_time(start)} --> {_srt_time(end)}\n")
            f.write(f"{en_text}\n")
            f.write(f"{cn_text}\n\n")
    print(f"✅ 字幕檔: {output_path}")


def write_description(topic: str, chapter_entries: list, output_path: str):
    """寫出含 YouTube 章節時間戳的影片描述"""
    lines = [
        f"🎓 {topic} - English Vocabulary Practice",
        f"Powered by {BRAND_NAME}",
        "",
        f"📚 Master essential {topic} vocabulary with IPA pronunciation,",
        "Chinese translation, and real-world example sentences.",
        "",
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
        "📑 CHAPTERS",
    ]
    for time_s, label in chapter_entries:
        lines.append(f"{_chapter_time(time_s)} {label}")
    lines += [
        "",
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
        "#EnglishLearning #Vocabulary #ESL #English #LearnEnglish",
        f"#{topic.replace(' ', '')}English",
    ]
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"✅ 影片描述: {output_path}")


def export_to_flashcard_app(data_list: list, topic: str, category: str = "videofactory") -> bool:
    """
    將詞彙卡片發布至 Flashcard App 公版素材庫（Explore）。
    寫入路徑：categories / packages / languageCards（全域公版，需 Service Account）。

    category:
      "videofactory" → 影片工廠分類（模式 2/3）
      "trending"     → 時事探索分類（模式 1，Tavily 知識卡）

    需要 .env 設定：
      FIREBASE_SERVICE_ACCOUNT=/path/to/serviceAccountKey.json
    """
    sa_path = os.getenv("FIREBASE_SERVICE_ACCOUNT")
    if not sa_path:
        print("⚠️  未設定 FIREBASE_SERVICE_ACCOUNT，跳過 Firestore 匯出")
        return False

    try:
        import firebase_admin
        from firebase_admin import credentials, firestore as fs
    except ImportError:
        print("⚠️  firebase-admin 未安裝，請執行: pip install firebase-admin")
        return False

    if not firebase_admin._apps:
        firebase_admin.initialize_app(credentials.Certificate(sa_path))
    db = fs.client()

    safe_topic = topic.lower().replace(" ", "-")

    # 分類設定
    if category == "trending":
        CATEGORY_ID  = "trending"
        category_doc = {"id": "trending", "name": {"en": "Trending", "zh": "時事探索"}, "order": 100}
        package_id   = f"trending-{safe_topic}"
    else:
        CATEGORY_ID  = "videofactory"
        category_doc = {"id": "videofactory", "name": {"en": "VideoFactory", "zh": "影片工廠"}, "order": 999}
        package_id   = f"videofactory-{safe_topic}"

    fs_batch = db.batch()

    # 1. 確保分類存在（merge=True 幂等）
    cat_ref = db.collection("categories").document(CATEGORY_ID)
    fs_batch.set(cat_ref, category_doc, merge=True)

    # 2. 建立（或覆寫）主題包
    pkg_ref = db.collection("packages").document(package_id)
    fs_batch.set(pkg_ref, {
        "id":         package_id,
        "categoryId": CATEGORY_ID,
        "name":       {"en": topic, "zh": topic},
        "order":      0,
        "cardCount":  len(data_list),
        "coverImage": "",
    })

    # 3. 寫入卡片（cardId 格式：{packageId}-{001}）
    for item in data_list:
        card_doc_id = f"{package_id}-{item['id']}"
        card_ref    = db.collection("languageCards").document(card_doc_id)
        fs_batch.set(card_ref, {
            "categoryId":       CATEGORY_ID,
            "packageId":        package_id,
            "cardId":           card_doc_id,
            "textEn":           item["word_en"],
            "textZhHant":       item["word_cn"],
            "textZhHans":       item["word_cn"],
            "ipaPronunciation": item["word_ipa"],
            "exampleEn":        item["sentence_en"],
            "exampleZhHant":    item["sentence_cn"],
            "exampleZhHans":    item["sentence_cn"],
            "difficulty":       2,
            "tags":             [safe_topic],
            "order":            int(item["id"]),
        })

    fs_batch.commit()
    print(f"✅ 已發布 {len(data_list)} 張卡片至 Explore 素材庫（包：{package_id}）")
    return True


# ================= 主流程 =================

def _flush_stdin() -> None:
    """排空 stdin 緩衝，避免網路等待期間誤按的 Enter 被後續 input() 消耗。"""
    if sys.stdin.isatty():
        try:
            termios.tcflush(sys.stdin, termios.TCIFLUSH)
        except Exception:
            pass
        # termios.tcflush 在部分 macOS 環境不可靠，用 select 二次確認排空
        try:
            import select as _select
            while _select.select([sys.stdin], [], [], 0)[0]:
                os.read(sys.stdin.fileno(), 4096)
        except Exception:
            pass


async def main():
    print("=" * 55)
    print("  🚀 VideoFactory Enterprise Edition")
    print("=" * 55)
    print()
    print("  請確認 .env 已設定以下金鑰：")
    print("    OPENAI_API_KEY           — 詞彙生成（必要）")
    print("    TAVILY_API_KEY           — Tavily 知識搜尋（來源 2 需要）")
    print("    PEXELS_API_KEY           — 背景圖片（格式 2/3 選用）")
    print("    FIREBASE_SERVICE_ACCOUNT — Firestore 匯出（格式 1/3 必要）")
    print()

    # ════════════════════════════════════════════════════
    # 輸入狀態機（step 0-4，q = 返回上一步）
    # ════════════════════════════════════════════════════
    step = 0
    user_input = context = source_type = topic = None
    count = 10
    fmt = 2
    range_str = ""
    intro_text = ""

    while step <= 5:
        if step == 0:
            print("  【第一階段】輸入內容來源（系統自動判斷模式）：")
            print("  · 主題文字      e.g.  Airport")
            print("  · Tavily 搜尋   e.g.  /search Apple Vision Pro  或  /s 川普關稅")
            print("  · 一般網址      e.g.  https://techcrunch.com/...")
            print("  · YouTube URL   e.g.  https://www.youtube.com/watch?v=...")
            print("  · 貼入長文      輸入  /text  後按 Enter，再貼上多行文字")
            print()
            raw = input("📌 請輸入: ").strip()
            if not raw:
                raw = "Airport"
            if raw.lower() == "/text":
                print("請貼上長文，完成後請在新的一行按下 Ctrl+D (Mac) 或 Ctrl+Z (Win) 結束：")
                lines = []
                try:
                    while True:
                        lines.append(input())
                except EOFError:
                    pass
                user_input = "\n".join(lines).strip()
                if not user_input:
                    print("⚠️  未輸入任何內容，返回主選單。")
                    continue
            else:
                user_input = raw
            context, source_type = route_input(user_input)
            _flush_stdin()  # 立即排空網路等待期間的誤按
            step = 1

        elif step == 1:
            if source_type in ("normal", "tavily"):
                # 自動計算 topic，不需詢問
                if source_type == "normal":
                    topic = user_input
                else:
                    lower = user_input.lower()
                    for prefix in _SEARCH_PREFIXES:
                        if lower.startswith(prefix):
                            topic = user_input[len(prefix):].strip()
                            break
                    else:
                        topic = user_input
                step = 2
            else:
                _flush_stdin()
                raw = input("📝 主題名稱（用於檔名與卡片包，e.g., AI Trends）(q=返回): ").strip()
                if raw.lower() == "q":
                    step = 0
                    continue
                topic = raw or "Topic"
                step = 2

        elif step == 2:
            _flush_stdin()
            raw = input("🔢 詞彙數量 (1-200，預設 10，q=返回): ").strip()
            if raw.lower() == "q":
                step = 0 if source_type in ("normal", "tavily") else 1
                continue
            count = int(raw) if raw.isdigit() else 10
            count = max(1, min(200, count))

            # Firestore 分類由 source_type 決定（與產出格式無關）
            fs_category = "videofactory" if source_type == "normal" else "trending"
            src_label = {"normal": "日常主題", "tavily": "Tavily 搜尋",
                         "youtube": "YouTube 字幕", "url": "網頁爬取", "text": "長文貼入"}
            print(f"\n   💡 來源模式：{src_label.get(source_type, source_type)}  |  Firestore → {fs_category}\n")
            step = 3

        elif step == 3:
            print("  【第二階段】選擇產出格式：")
            print("  [1] 純產字卡 (Cards Only)  — LLM → Firestore，不產影片")
            print("  [2] 純產影片 (Video Only)  — LLM → MP4，不寫入 Firestore")
            print("  [3] 影卡雙棲 (Both)        — LLM → MP4 + Firestore")
            _flush_stdin()
            raw = input("\n  選擇格式 (1/2/3，預設 2，q=返回): ").strip()
            if raw.lower() == "q":
                step = 2
                continue
            fmt = int(raw) if raw in ("1", "2", "3") else 2
            print()
            step = 4

        elif step == 4:
            if fmt == 1:
                break
            _flush_stdin()
            raw = input("📐 處理範圍 (選填，格式 3-40，留空=全部，q=返回): ").strip()
            if raw.lower() == "q":
                step = 3
                continue
            range_str = raw
            step = 5

        elif step == 5:
            _flush_stdin()
            print("🎙️  【片頭語音】輸入片頭旁白文字（留空=使用原始 prebuilt intro，q=返回）：")
            raw = input("   文字: ").strip()
            if raw.lower() == "q":
                step = 4
                continue
            intro_text = raw
            break

    # ════════════════════════════════════════════════════
    # 格式 1：純產字卡（Cards Only）
    # ════════════════════════════════════════════════════
    if fmt == 1:
        if not OPENAI_KEYS:
            print("❌ 需要 OPENAI_API_KEY")
            return
        try:
            data_list = generate_content(topic, count, context=context)
            with open(DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(data_list, f, ensure_ascii=False, indent=2)
            print(f"✅ 已生成 {len(data_list)} 個詞彙")
        except Exception as e:
            print(f"❌ OpenAI 生成失敗: {e}")
            return

        if not data_list:
            print("❌ 沒有可處理的教學內容！")
            return

        export_to_flashcard_app(data_list, topic, category=fs_category)

        safe_slug = topic.lower().replace(" ", "-")
        pkg_id    = f"{fs_category}-{safe_slug}"
        print(f"""
{'=' * 55}
  🎉 字卡已發布！
{'=' * 55}
  📦 Package  : {pkg_id}
  🗂  Category : {fs_category}
  🃏 卡片數量  : {len(data_list)} 張
{'=' * 55}
""")
        return

    # ════════════════════════════════════════════════════
    # 格式 2 / 3：影片流程
    # ════════════════════════════════════════════════════
    fmt_label = "純產影片" if fmt == 2 else "影卡雙棲"
    print(f"📋 格式：{fmt_label}\n")

    # 範圍解析（range_str 由狀態機收集）
    range_start, range_end = None, None
    if range_str and "-" in range_str:
        parts = range_str.split("-", 1)
        try:
            range_start = int(parts[0].strip())
            range_end   = int(parts[1].strip())
        except ValueError:
            print("⚠️  範圍格式錯誤，將處理全部")
    print()

    # ── 素材檢查 ─────────────────────────────────────
    if not check_assets():
        return

    # ── 生成教學內容 ──────────────────────────────────
    if OPENAI_KEYS:
        try:
            data_list = generate_content(topic, count, context=context)
            with open(DATA_FILE, "w", encoding="utf-8") as f:
                json.dump(data_list, f, ensure_ascii=False, indent=2)
            print(f"✅ 已生成 {len(data_list)} 個詞彙並儲存至 data.json")

            # ── 內容審核關卡 ──────────────────────────────────
            review_path = export_review_excel(data_list, topic)
            print(f"\n📊 內容審核 Excel：{review_path}")
            try:
                if sys.platform == "win32":
                    os.startfile(review_path)
                elif sys.platform == "darwin":
                    subprocess.run(["open", review_path], check=True)
                else:
                    subprocess.run(["xdg-open", review_path], check=True)
            except Exception:
                pass  # 無法自動開啟也繼續

            _confirm = input("\n✅ 請在 Excel 確認（或修改）內容後按 Enter 繼續，輸入 q 中止：").strip().lower()
            if _confirm == "q":
                print("⛔ 已中止。")
                return

            # 重新從 Excel 讀回（自動套用使用者對 Excel 的修改）
            data_list = import_review_excel(review_path)
            with open(DATA_FILE, "w", encoding="utf-8") as _f:
                json.dump(data_list, _f, ensure_ascii=False, indent=2)
            print(f"✅ 已讀回 {len(data_list)} 個詞彙（含任何修改）\n")

        except Exception as e:
            print(f"⚠️  OpenAI 失敗（{type(e).__name__}: {e}），嘗試讀取 data.json ...")
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data_list = json.load(f)
    else:
        print("⚠️  未設定任何 OPENAI_API_KEY，直接讀取 data.json")
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            data_list = json.load(f)

    if not data_list:
        print("❌ 沒有可處理的教學內容！")
        return

    # 套用範圍篩選（1-based，包含頭尾）
    total = len(data_list)
    if range_start is not None and range_end is not None:
        s = max(1, range_start) - 1
        e = min(total, range_end)
        data_list = data_list[s:e]
        print(f"📐 處理範圍：第 {range_start} 張 ～ 第 {range_end} 張，共 {len(data_list)} 張")
    else:
        print(f"📋 共 {total} 個詞彙，全部處理")

    print(f"   實際處理：{len(data_list)} 張\n")

    # ── 4. 下載 Pexels 圖片 ─────────────────────────────
    # <=10 張：下載相同張數；11-20 張：下載 10 張隨機用；>20 張：下載 ceil(count/2) 張
    vocab_count = len(data_list)
    if vocab_count <= 10:
        pexels_count = vocab_count
    elif vocab_count <= 20:
        pexels_count = 10
    else:
        pexels_count = math.ceil(vocab_count / 2)

    pexels_images: list = []
    if PEXELS_KEY:
        try:
            pexels_images = download_pexels_images(topic, count=pexels_count)
        except Exception as e:
            print(f"⚠️  Pexels 下載失敗 ({e})，將使用本地圖片")
    else:
        print("⚠️  未設定 PEXELS_API_KEY，使用 assets/bg/ 本地圖片\n")

    # ── 5. 客製片頭（若有輸入文字）──────────────────────
    active_intro = INTRO_VIDEO
    if intro_text:
        custom_path = await generate_custom_intro(intro_text)
        if custom_path:
            active_intro = custom_path
            print(f"✅  使用客製片頭：{os.path.basename(custom_path)}")
        else:
            print("↩️  取消客製，改用預設 prebuilt intro")

    # ── 6. 雙階段結構 ───────────────────────────────────
    srt_entries:     list  = []
    chapter_entries: list  = []
    cumulative_time: float = 0.0

    if os.path.exists(active_intro):
        dur = _video_duration(active_intro)
        chapter_entries.append((0.0, "Intro"))
        cumulative_time += dur
        print(f"🎬 Intro 已偵測 ({dur:.1f}s)")

    # ── 6. Phase 1：完整教學 ────────────────────────────
    print(f"\n▶ Phase 1（完整教學）：{len(data_list)} 個詞彙")
    chunks_g0, cumulative_time = await process_group(
        data_list, 0, pexels_images, cumulative_time, srt_entries, chapter_entries,
        phase=1,
    )

    if os.path.exists(BREAK_VIDEO):
        dur = _video_duration(BREAK_VIDEO)
        chapter_entries.append((cumulative_time, "⏸ Active Recall Challenge"))
        cumulative_time += dur
        print(f"\n⏸ Break 已偵測 ({dur:.1f}s)")

    # ── 7. Phase 2：Active Recall ───────────────────────
    print(f"\n▶ Phase 2（Active Recall）：{len(data_list)} 個詞彙")
    chunks_g1, cumulative_time = await process_group(
        data_list, 1, pexels_images, cumulative_time, srt_entries, chapter_entries,
        phase=2,
    )

    if os.path.exists(OUTRO_VIDEO):
        chapter_entries.append((cumulative_time, "Outro"))

    # ── 8. FFmpeg 高速合併所有片段 ──────────────────────
    print("\n🎬 正在使用 FFmpeg 組裝最終影片...")

    safe_topic   = topic.lower().replace(" ", "_")
    concat_path  = os.path.join(TEMP_DIR, "concat_list.txt")
    merged_path  = os.path.join(TEMP_DIR, "merged_no_bgm.mp4")
    output_file  = os.path.join(OUTPUT_DIR, f"final_{safe_topic}.mp4")
    srt_path     = os.path.join(OUTPUT_DIR, f"final_{safe_topic}.srt")
    desc_path    = os.path.join(OUTPUT_DIR, f"description_{safe_topic}.txt")

    with open(concat_path, "w", encoding="utf-8") as f:
        if os.path.exists(active_intro):
            f.write(f"file '{_normalize_prebuilt(active_intro)}'\n")
        for chunk in chunks_g0:
            f.write(f"file '{chunk}'\n")
        if os.path.exists(BREAK_VIDEO):
            f.write(f"file '{_normalize_prebuilt(BREAK_VIDEO)}'\n")
        for chunk in chunks_g1:
            f.write(f"file '{chunk}'\n")
        if os.path.exists(OUTRO_VIDEO):
            f.write(f"file '{_normalize_prebuilt(OUTRO_VIDEO)}'\n")

    ret = os.system(
        f'ffmpeg -y -f concat -safe 0 -i "{concat_path}" '
        f'-c copy "{merged_path}" -loglevel error'
    )
    if ret != 0 or not os.path.exists(merged_path):
        print("⚠️  FFmpeg concat 失敗，嘗試完整重新編碼...")
        os.system(
            f'ffmpeg -y -f concat -safe 0 -i "{concat_path}" '
            f'-c:v libx264 -preset fast -c:a aac -ar 48000 -ac 2 "{merged_path}" -loglevel error'
        )

    # ── 9. 輸出（無背景音樂）─────────────────────────────
    shutil.copy(merged_path, output_file)

    # ── 10. 輸出 SRT + 影片描述 ────────────────────────
    write_srt(srt_entries, srt_path)
    write_description(topic, chapter_entries, desc_path)

    # ── 11. 格式 3：發布至 Firestore（分類由來源決定）──
    if fmt == 3:
        export_to_flashcard_app(data_list, topic, category=fs_category)

    print(f"""
{'=' * 55}
  🎉 製作完成！
{'=' * 55}
  🎬 影片  : {output_file}
  📝 字幕  : {srt_path}
  📄 描述  : {desc_path}
{'=' * 55}
""")


if __name__ == "__main__":
    asyncio.run(main())
